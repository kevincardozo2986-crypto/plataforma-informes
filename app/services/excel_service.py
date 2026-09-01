"""Creación progresiva y consulta del libro Excel de trabajo."""

import re
import gc
import shutil
import tempfile
from collections import Counter, defaultdict
from statistics import median
from pathlib import Path

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.chart_service import (
    agregar_docentes_dg_xlsxwriter,
    agregar_estudiantes_dg2_xlsxwriter,
    agregar_estudiantes_dg_xlsxwriter,
)


SHEET_NAMES = (
    "Original",
    "Tabla Dinamica Docentes",
    "Docentes DG",
    "Tabla Dinamica Estudiantes",
    "Estudiantes DG",
    "Estudiantes DG2",
    "Tabla Dinamica Actividades",
    "Resumen Informe",
    "Diseño de Cursos",
)

MESES_ABREVIADOS = {
    1: "ENE",
    2: "FEB",
    3: "MAR",
    4: "ABR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AGO",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DIC",
}


class ExcelProcess:
    """Mantiene un único archivo temporal durante todo el proceso."""

    def __init__(self, existing_path=None):
        self._temporary_directory = None
        if existing_path:
            self.path = Path(existing_path)
            self.path.parent.mkdir(parents=True, exist_ok=True)
        else:
            self._temporary_directory = tempfile.TemporaryDirectory(
                prefix="plataforma_informes_"
            )
            self.path = Path(self._temporary_directory.name) / "informe_en_proceso.xlsx"
        self._row_counts = {}
        self._teacher_summary_cache = None

    @property
    def exists(self):
        return self.path.is_file()

    def create_original(self, datos):
        self._write_original(datos)

    def create_original_from_chunks(self, chunks, total_rows=None, progress_callback=None):
        """Crea Original con escritura rápida y memoria constante."""
        libro_excel = xlsxwriter.Workbook(
            self.path,
            {
                "constant_memory": True,
                "default_date_format": "dd/mm/yyyy",
                "strings_to_urls": False,
            },
        )
        hoja_original = libro_excel.add_worksheet("Original")
        formato_encabezado = libro_excel.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center"}
        )
        cantidad_filas = 0
        cantidad_columnas = 0
        indices_docentes = None
        dias_por_mes_cache = defaultdict(lambda: defaultdict(set))
        dias_periodo_cache = defaultdict(set)
        try:
            for bloque_datos in chunks:
                if cantidad_columnas == 0:
                    encabezados = [str(valor) for valor in bloque_datos.columns]
                    cantidad_columnas = len(encabezados)
                    columnas_normalizadas = {
                        encabezado.strip().casefold(): indice
                        for indice, encabezado in enumerate(encabezados)
                    }
                    requeridas_docentes = ("rol", "curso", "usuario", "mes", "dia")
                    if all(
                        nombre in columnas_normalizadas
                        for nombre in requeridas_docentes
                    ):
                        indices_docentes = {
                            nombre: columnas_normalizadas[nombre]
                            for nombre in requeridas_docentes
                        }
                    hoja_original.write_row(0, 0, encabezados, formato_encabezado)
                    for indice, encabezado in enumerate(encabezados):
                        hoja_original.set_column(indice, indice, min(max(len(encabezado) + 2, 11), 30))
                for valores_fila in bloque_datos.itertuples(index=False, name=None):
                    hoja_original.write_row(
                        cantidad_filas + 1,
                        0,
                        [self._excel_value(valor) for valor in valores_fila],
                    )
                    if indices_docentes:
                        self._acumular_actividad_docente(
                            valores_fila,
                            indices_docentes,
                            dias_por_mes_cache,
                            dias_periodo_cache,
                        )
                    cantidad_filas += 1
                if progress_callback and total_rows:
                    progress_callback(min(int(cantidad_filas * 96 / total_rows) + 2, 98))
            if cantidad_columnas == 0:
                raise ValueError("El CSV no contiene columnas ni registros legibles.")
            hoja_original.freeze_panes(1, 0)
            hoja_original.autofilter(0, 0, cantidad_filas, cantidad_columnas - 1)
        finally:
            libro_excel.close()
        self._row_counts["Original"] = cantidad_filas
        self._teacher_summary_cache = (
            (dias_por_mes_cache, dias_periodo_cache)
            if dias_por_mes_cache
            else None
        )
        if progress_callback:
            progress_callback(100)
        return cantidad_filas, cantidad_columnas

    @staticmethod
    def _acumular_actividad_docente(
        valores_fila, indices, dias_por_mes, dias_periodo
    ):
        """Acumula solo los datos necesarios mientras Original ya se está escribiendo."""
        rol = valores_fila[indices["rol"]]
        if str(rol or "").strip().casefold() != "editingteacher":
            return
        curso = str(valores_fila[indices["curso"]] or "").strip()
        docente = str(valores_fila[indices["usuario"]] or "").strip()
        try:
            mes = int(float(valores_fila[indices["mes"]]))
            dia = int(float(valores_fila[indices["dia"]]))
        except (TypeError, ValueError):
            return
        if not curso or not docente or not 1 <= mes <= 12 or not 1 <= dia <= 31:
            return
        clave = (curso, docente)
        dias_por_mes[clave][mes].add(dia)
        dias_periodo[clave].add(dia)

    @staticmethod
    def _excel_value(value):
        if pd.isna(value):
            return None
        if hasattr(value, "to_pydatetime"):
            value = value.to_pydatetime()
        elif hasattr(value, "item"):
            value = value.item()
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    def update_original(self, datos):
        self._write_original(datos)

    def _write_original(self, datos):
        modo_escritura = "a" if self.exists else "w"
        opciones_escritura = {"mode": modo_escritura, "engine": "openpyxl"}
        if modo_escritura == "a":
            opciones_escritura["if_sheet_exists"] = "replace"
        with pd.ExcelWriter(self.path, **opciones_escritura) as escritor:
            datos.to_excel(escritor, sheet_name="Original", index=False)
        if modo_escritura == "a":
            libro_excel = load_workbook(self.path)
            if "Tabla Dinamica Docentes" in libro_excel.sheetnames:
                libro_excel.remove(libro_excel["Tabla Dinamica Docentes"])
                libro_excel.save(self.path)
        self._row_counts["Original"] = len(datos)
        self._row_counts.pop("Tabla Dinamica Docentes", None)
        self._format_sheet("Original")

    def _format_sheet(self, sheet_name):
        libro_excel = load_workbook(self.path)
        hoja = libro_excel[sheet_name]
        relleno_encabezado = PatternFill("solid", fgColor="0A3A6B")
        for celda in hoja[1]:
            celda.fill = relleno_encabezado
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center")
        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = hoja.dimensions
        for indice, columna in enumerate(hoja.iter_cols(min_row=1, max_row=min(hoja.max_row, 200)), 1):
            ancho = max((len(str(celda.value or "")) for celda in columna), default=8)
            hoja.column_dimensions[get_column_letter(indice)].width = min(max(ancho + 2, 11), 38)
        libro_excel.save(self.path)

    def crear_tabla_docentes(self, progress_callback=None):
        """Crea la segunda hoja con días distintos por curso, docente y mes."""
        if not self.exists:
            raise FileNotFoundError("Primero debes crear la hoja Original.")
        if progress_callback:
            progress_callback(5)

        if self._teacher_summary_cache:
            dias_por_mes, dias_periodo = self._teacher_summary_cache
            if progress_callback:
                progress_callback(58)
            return self._guardar_tabla_docentes(
                dias_por_mes, dias_periodo, progress_callback
            )

        libro_lectura = load_workbook(self.path, read_only=True, data_only=True)
        hoja_original = libro_lectura["Original"]
        filas_originales = hoja_original.iter_rows(values_only=True)
        encabezados_originales = next(filas_originales, None)
        if not encabezados_originales:
            libro_lectura.close()
            raise ValueError("La hoja Original está vacía.")

        columnas = {
            str(nombre).strip().casefold(): indice
            for indice, nombre in enumerate(encabezados_originales)
        }
        requeridas = ("rol", "curso", "usuario", "mes", "dia")
        faltantes = [nombre for nombre in requeridas if nombre not in columnas]
        if faltantes:
            libro_lectura.close()
            raise ValueError(
                "La hoja Original no contiene las columnas requeridas: "
                + ", ".join(faltantes)
            )

        dias_por_mes = defaultdict(lambda: defaultdict(set))
        dias_periodo = defaultdict(set)
        total_filas_original = max((hoja_original.max_row or 1) - 1, 1)
        try:
            for numero_fila, fila in enumerate(filas_originales, 1):
                rol = fila[columnas["rol"]]
                if str(rol or "").strip().casefold() != "editingteacher":
                    if progress_callback and numero_fila % 5_000 == 0:
                        progress_callback(
                            min(5 + int(numero_fila * 50 / total_filas_original), 55)
                        )
                    continue

                curso = str(fila[columnas["curso"]] or "").strip()
                docente = str(fila[columnas["usuario"]] or "").strip()
                try:
                    mes = int(float(fila[columnas["mes"]]))
                    dia = int(float(fila[columnas["dia"]]))
                except (TypeError, ValueError):
                    continue
                if not curso or not docente or not 1 <= mes <= 12 or not 1 <= dia <= 31:
                    continue

                clave = (curso, docente)
                dias_por_mes[clave][mes].add(dia)
                dias_periodo[clave].add(dia)
                if progress_callback and numero_fila % 5_000 == 0:
                    progress_callback(
                        min(5 + int(numero_fila * 50 / total_filas_original), 55)
                    )
        finally:
            libro_lectura.close()

        if not dias_por_mes:
            raise ValueError(
                "No se encontraron registros con rol 'editingteacher' en Original."
            )
        if progress_callback:
            progress_callback(58)
        return self._guardar_tabla_docentes(
            dias_por_mes, dias_periodo, progress_callback
        )

    def _guardar_tabla_docentes(
        self, dias_por_mes, dias_periodo, progress_callback=None
    ):
        """Escribe la tabla pequeña usando el resumen acumulado durante Original."""
        meses_encontrados = sorted(
            {mes for meses in dias_por_mes.values() for mes in meses}
        )
        encabezados = [
            "CURSO",
            "DOCENTE",
            *(MESES_ABREVIADOS[mes] for mes in meses_encontrados),
            "TOTAL",
        ]
        filas = []
        for clave in sorted(
            dias_por_mes,
            key=lambda valores: (valores[0].casefold(), valores[1].casefold()),
        ):
            curso, docente = clave
            filas.append(
                [
                    curso,
                    docente,
                    *(len(dias_por_mes[clave][mes]) for mes in meses_encontrados),
                ]
            )
            filas[-1].append(sum(filas[-1][2:]))
        promedios_mensuales = [
            round(sum(fila[indice + 2] for fila in filas) / len(filas), 1)
            for indice in range(len(meses_encontrados))
        ]
        nombre_hoja = "Tabla Dinamica Docentes"
        ruta_nueva = self.path.with_name("informe_actualizado.xlsx")
        libro_lectura = load_workbook(self.path, read_only=True, data_only=False)
        hoja_lectura = libro_lectura["Original"]
        total_original = max(hoja_lectura.max_row or 1, 1)
        libro_salida = xlsxwriter.Workbook(
            ruta_nueva,
            {
                "constant_memory": True,
                "default_date_format": "dd/mm/yyyy",
                "strings_to_urls": False,
            },
        )
        formato_encabezado = libro_salida.add_format(
            {
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#0A3A6B",
                "align": "center",
                "valign": "vcenter",
            }
        )
        formato_centrado = libro_salida.add_format(
            {"align": "center", "valign": "vcenter"}
        )
        formato_total = libro_salida.add_format(
            {
                "bold": True,
                "bg_color": "#E7F1FA",
                "font_color": "#0A3A6B",
                "align": "center",
                "valign": "vcenter",
            }
        )
        formato_promedio = libro_salida.add_format(
            {
                "bold": True,
                "bg_color": "#DDEAF6",
                "font_color": "#0A3A6B",
                "align": "center",
                "valign": "vcenter",
                "top": 1,
                "top_color": "#0A4D91",
                "num_format": "0.0",
            }
        )
        try:
            original_salida = libro_salida.add_worksheet("Original")
            for indice_fila, valores in enumerate(
                hoja_lectura.iter_rows(values_only=True)
            ):
                formato = formato_encabezado if indice_fila == 0 else None
                original_salida.write_row(
                    indice_fila,
                    0,
                    [self._excel_value(valor) for valor in valores],
                    formato,
                )
                if indice_fila == 0:
                    for indice_columna, encabezado in enumerate(valores):
                        ancho = min(max(len(str(encabezado or "")) + 2, 11), 30)
                        original_salida.set_column(
                            indice_columna, indice_columna, ancho
                        )
                if progress_callback and indice_fila % 5_000 == 0:
                    progress_callback(
                        min(60 + int(indice_fila * 28 / total_original), 88)
                    )
            original_salida.freeze_panes(1, 0)
            if hoja_lectura.max_column:
                original_salida.autofilter(
                    0, 0, max(total_original - 1, 0), hoja_lectura.max_column - 1
                )

            tabla_salida = libro_salida.add_worksheet(nombre_hoja)
            tabla_salida.write_row(0, 0, encabezados, formato_encabezado)
            for indice_fila, fila in enumerate(filas, 1):
                tabla_salida.write(indice_fila, 0, fila[0])
                tabla_salida.write(indice_fila, 1, fila[1])
                tabla_salida.write_row(
                    indice_fila, 2, fila[2:-1], formato_centrado
                )
                tabla_salida.write(
                    indice_fila, len(encabezados) - 1, fila[-1], formato_total
                )
            fila_promedio = len(filas) + 1
            tabla_salida.write(fila_promedio, 0, "TOTAL GENERAL", formato_promedio)
            tabla_salida.write(fila_promedio, 1, "", formato_promedio)
            tabla_salida.write_row(
                fila_promedio, 2, promedios_mensuales, formato_promedio
            )
            tabla_salida.write(
                fila_promedio, len(encabezados) - 1, "", formato_promedio
            )
            tabla_salida.set_column(0, 0, 42)
            tabla_salida.set_column(1, 1, 34)
            tabla_salida.set_column(2, len(encabezados) - 1, 14)
            tabla_salida.freeze_panes(1, 2)
            tabla_salida.autofilter(0, 0, len(filas), len(encabezados) - 1)
            if progress_callback:
                progress_callback(92)
        finally:
            libro_lectura.close()
            libro_salida.close()

        ruta_nueva.replace(self.path)

        cantidad_filas = len(filas)
        self._row_counts[nombre_hoja] = cantidad_filas + 1
        if progress_callback:
            progress_callback(100)
        return cantidad_filas, [MESES_ABREVIADOS[mes] for mes in meses_encontrados]

    def sheet_names(self):
        if not self.exists:
            return []
        libro_excel = load_workbook(self.path, read_only=True)
        nombres_hojas = list(libro_excel.sheetnames)
        libro_excel.close()
        return nombres_hojas

    def crear_grafica_docentes(self, programa, periodo, progress_callback=None):
        """Agrega Docentes DG copiando el libro por streaming para archivos grandes."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        if progress_callback:
            progress_callback(3)
        libro_lectura = load_workbook(self.path, read_only=True, data_only=True)
        if "Tabla Dinamica Docentes" not in libro_lectura.sheetnames:
            libro_lectura.close()
            raise ValueError(
                "No existe la hoja 'Tabla Dinamica Docentes'. Créala antes de generar la gráfica."
            )

        tabla = libro_lectura["Tabla Dinamica Docentes"]
        encabezados = [str(celda.value or "").strip() for celda in tabla[1]]
        meses_validos = set(MESES_ABREVIADOS.values())
        columnas_mensuales = [
            (indice, nombre.upper())
            for indice, nombre in enumerate(encabezados)
            if nombre.upper() in meses_validos
        ]
        if not columnas_mensuales:
            libro_lectura.close()
            raise ValueError(
                "La hoja 'Tabla Dinamica Docentes' no contiene columnas mensuales."
            )
        valores_por_mes = {mes: [] for _, mes in columnas_mensuales}
        for fila in tabla.iter_rows(min_row=2, values_only=True):
            etiqueta = str(fila[0] or "").strip().upper()
            if etiqueta in {"PROMEDIO", "TOTAL GENERAL"}:
                continue
            for indice, mes in columnas_mensuales:
                valor = fila[indice] if indice < len(fila) else None
                if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    valores_por_mes[mes].append(float(valor))
        promedios = [
            (mes, round(sum(valores) / len(valores), 1))
            for mes, valores in valores_por_mes.items()
            if valores
        ]
        if not promedios:
            libro_lectura.close()
            raise ValueError(
                "Las columnas mensuales no contienen valores numéricos para calcular promedios."
            )

        ruta_nueva = self.path.with_name("informe_con_grafica.xlsx")
        libro_salida = xlsxwriter.Workbook(
            ruta_nueva,
            {"constant_memory": True, "strings_to_urls": False},
        )
        formato_encabezado = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center"}
        )
        formato_centrado = libro_salida.add_format(
            {"align": "center", "valign": "vcenter"}
        )
        formato_total = libro_salida.add_format(
            {
                "bold": True, "bg_color": "#E7F1FA",
                "font_color": "#0A3A6B", "align": "center",
            }
        )
        formato_promedio = libro_salida.add_format(
            {
                "bold": True, "bg_color": "#DDEAF6",
                "font_color": "#0A3A6B", "align": "center",
                "top": 1, "top_color": "#0A4D91", "num_format": "0.0",
            }
        )
        total_filas = sum(
            max(libro_lectura[nombre].max_row or 1, 1)
            for nombre in libro_lectura.sheetnames
            if nombre != "Docentes DG"
        )
        copiadas = 0
        try:
            for nombre in libro_lectura.sheetnames:
                if nombre == "Docentes DG":
                    continue
                origen = libro_lectura[nombre]
                destino = libro_salida.add_worksheet(nombre)
                for indice_fila, fila in enumerate(origen.iter_rows(values_only=True)):
                    valores = [self._excel_value(valor) for valor in fila]
                    if indice_fila == 0:
                        destino.write_row(indice_fila, 0, valores, formato_encabezado)
                    elif nombre == "Tabla Dinamica Docentes":
                        es_promedio = str(valores[0] or "").strip().upper() in {
                            "PROMEDIO", "TOTAL GENERAL"
                        }
                        if es_promedio:
                            destino.write_row(indice_fila, 0, valores, formato_promedio)
                        else:
                            destino.write(indice_fila, 0, valores[0])
                            destino.write(indice_fila, 1, valores[1])
                            destino.write_row(indice_fila, 2, valores[2:-1], formato_centrado)
                            destino.write(indice_fila, len(valores) - 1, valores[-1], formato_total)
                    else:
                        destino.write_row(indice_fila, 0, valores)
                    copiadas += 1
                    if progress_callback and copiadas % 2_000 == 0:
                        progress_callback(min(5 + int(copiadas * 85 / total_filas), 90))
                destino.freeze_panes(1, 2 if nombre == "Tabla Dinamica Docentes" else 0)
                if nombre == "Original" and origen.max_column:
                    destino.autofilter(0, 0, max((origen.max_row or 1) - 1, 0), origen.max_column - 1)
                    for indice, encabezado in enumerate(encabezados if nombre == "Tabla Dinamica Docentes" else [c.value for c in origen[1]]):
                        ancho = min(max(len(str(encabezado or "")) + 2, 11), 30)
                        destino.set_column(indice, indice, ancho)
                elif nombre == "Tabla Dinamica Docentes":
                    destino.set_column(0, 0, 42)
                    destino.set_column(1, 1, 34)
                    destino.set_column(2, max(origen.max_column - 1, 2), 14)
                    destino.autofilter(0, 0, max((origen.max_row or 2) - 2, 0), origen.max_column - 1)
            agregar_docentes_dg_xlsxwriter(
                libro_salida, promedios, programa, periodo
            )
            if progress_callback:
                progress_callback(94)
        finally:
            libro_lectura.close()
            libro_salida.close()
        ruta_nueva.replace(self.path)
        self._row_counts["Docentes DG"] = len(promedios)
        if progress_callback:
            progress_callback(100)
        return promedios

    def crear_tabla_estudiantes(self, programa, periodo, progress_callback=None):
        """Crea días y estudiantes únicos por curso/mes leyendo Original por streaming."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        libro_lectura = load_workbook(self.path, read_only=True, data_only=True)
        if "Original" not in libro_lectura.sheetnames:
            libro_lectura.close()
            raise ValueError("No existe la hoja 'Original'.")
        original = libro_lectura["Original"]
        filas = original.iter_rows(values_only=True)
        encabezados = next(filas, None)
        if not encabezados:
            libro_lectura.close()
            raise ValueError("La hoja Original está vacía.")
        columnas = {
            str(nombre or "").strip().casefold(): indice
            for indice, nombre in enumerate(encabezados)
        }
        requeridas = ("rol", "curso", "mes", "dia", "idusuario")
        faltantes = [nombre for nombre in requeridas if nombre not in columnas]
        if faltantes:
            libro_lectura.close()
            raise ValueError(
                "La hoja Original no contiene las columnas requeridas: "
                + ", ".join(faltantes)
            )

        dias = defaultdict(lambda: defaultdict(set))
        estudiantes = defaultdict(lambda: defaultdict(set))
        total_original = max((original.max_row or 1) - 1, 1)
        for numero, fila in enumerate(filas, 1):
            if str(fila[columnas["rol"]] or "").strip().casefold() != "student":
                continue
            curso = str(fila[columnas["curso"]] or "").strip()
            usuario = fila[columnas["idusuario"]]
            try:
                mes = int(float(fila[columnas["mes"]]))
                dia = int(float(fila[columnas["dia"]]))
            except (TypeError, ValueError):
                continue
            if not curso or usuario in (None, "") or not 1 <= mes <= 12 or not 1 <= dia <= 31:
                continue
            dias[curso][mes].add(dia)
            estudiantes[curso][mes].add(str(usuario).strip())
            if progress_callback and numero % 5_000 == 0:
                progress_callback(min(5 + int(numero * 42 / total_original), 47))
        if not dias:
            libro_lectura.close()
            raise ValueError("No se encontraron registros con rol 'student' en Original.")

        meses = sorted({mes for cursos in dias.values() for mes in cursos})
        cursos = sorted(dias, key=str.casefold)
        ruta_nueva = self.path.with_name("informe_con_estudiantes.xlsx")
        libro_salida = xlsxwriter.Workbook(
            ruta_nueva, {"constant_memory": True, "strings_to_urls": False}
        )
        formato_encabezado = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center", "valign": "vcenter"}
        )
        formato_subencabezado = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#1767A6", "align": "center", "valign": "vcenter"}
        )
        formato_mes = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center_across", "valign": "vcenter"}
        )
        formato_numero = libro_salida.add_format({"align": "center", "valign": "vcenter"})
        formato_total = libro_salida.add_format(
            {"bold": True, "bg_color": "#E7F1FA", "font_color": "#0A3A6B", "align": "center"}
        )
        formato_resumen = libro_salida.add_format(
            {"bold": True, "bg_color": "#DDEAF6", "font_color": "#0A3A6B", "align": "center", "top": 1, "num_format": "0.0"}
        )
        hojas_copiadas = [
            n for n in libro_lectura.sheetnames
            if n not in {"Tabla Dinamica Estudiantes", "Estudiantes DG", "Estudiantes DG2"}
        ]
        total_copia = sum(max(libro_lectura[n].max_row or 1, 1) for n in hojas_copiadas)
        copiadas = 0
        try:
            for nombre in hojas_copiadas:
                origen = libro_lectura[nombre]
                if nombre == "Docentes DG":
                    valores_dg = [
                        (str(fila[0]), float(fila[1]))
                        for fila in origen.iter_rows(min_row=2, values_only=True)
                        if fila[0] not in (None, "") and isinstance(fila[1], (int, float))
                    ]
                    agregar_docentes_dg_xlsxwriter(libro_salida, valores_dg, programa, periodo)
                    copiadas += max(origen.max_row or 1, 1)
                    continue
                destino = libro_salida.add_worksheet(nombre)
                for indice_fila, valores in enumerate(origen.iter_rows(values_only=True)):
                    valores_limpios = [self._excel_value(valor) for valor in valores]
                    if indice_fila == 0:
                        destino.write_row(indice_fila, 0, valores_limpios, formato_encabezado)
                    elif nombre == "Tabla Dinamica Docentes":
                        es_resumen = str(valores_limpios[0] or "").strip().upper() in {
                            "PROMEDIO", "TOTAL GENERAL"
                        }
                        if es_resumen:
                            destino.write_row(indice_fila, 0, valores_limpios, formato_resumen)
                        else:
                            destino.write(indice_fila, 0, valores_limpios[0])
                            destino.write(indice_fila, 1, valores_limpios[1])
                            destino.write_row(indice_fila, 2, valores_limpios[2:-1], formato_numero)
                            destino.write(indice_fila, len(valores_limpios) - 1, valores_limpios[-1], formato_total)
                    else:
                        destino.write_row(indice_fila, 0, valores_limpios)
                    copiadas += 1
                    if progress_callback and copiadas % 2_000 == 0:
                        progress_callback(min(48 + int(copiadas * 40 / total_copia), 88))
                destino.freeze_panes(1, 2 if nombre == "Tabla Dinamica Docentes" else 0)
                if nombre == "Original" and origen.max_column:
                    destino.autofilter(0, 0, max((origen.max_row or 1) - 1, 0), origen.max_column - 1)
                    for indice, encabezado in enumerate(encabezados):
                        ancho = min(max(len(str(encabezado or "")) + 2, 11), 30)
                        destino.set_column(indice, indice, ancho)
                elif nombre == "Tabla Dinamica Docentes":
                    destino.set_column(0, 0, 42)
                    destino.set_column(1, 1, 34)
                    destino.set_column(2, max(origen.max_column - 1, 2), 14)
                    destino.autofilter(0, 0, max((origen.max_row or 2) - 2, 0), origen.max_column - 1)

            hoja = libro_salida.add_worksheet("Tabla Dinamica Estudiantes")
            cantidad_meses = len(meses)
            columna_dias = 1
            columna_total_dias = columna_dias + cantidad_meses
            columna_estudiantes = columna_total_dias + 1
            columna_total_estudiantes = columna_estudiantes + cantidad_meses
            columna_indicador = columna_total_estudiantes + 2
            fila_datos = 3
            fila_total = fila_datos + len(cursos)
            fila_indicadores = fila_total + 3
            formato_decimal = libro_salida.add_format(
                {"align": "center", "valign": "vcenter", "num_format": "0.0"}
            )

            hoja.write(0, 0, "rol")
            hoja.write(0, 1, "student")
            hoja.write(1, 0, "CURSO", formato_encabezado)
            hoja.write(1, columna_dias, "DÍAS", formato_mes)
            for columna in range(columna_dias + 1, columna_total_dias + 1):
                hoja.write_blank(1, columna, None, formato_mes)
            hoja.write(1, columna_estudiantes, "ESTUDIANTES", formato_mes)
            for columna in range(columna_estudiantes + 1, columna_total_estudiantes + 1):
                hoja.write_blank(1, columna, None, formato_mes)
            hoja.write_blank(2, 0, None, formato_subencabezado)
            for posicion, mes in enumerate(meses):
                hoja.write(2, columna_dias + posicion, MESES_ABREVIADOS[mes], formato_subencabezado)
            hoja.write(2, columna_total_dias, "TOTAL", formato_subencabezado)
            for posicion, mes in enumerate(meses):
                hoja.write(2, columna_estudiantes + posicion, MESES_ABREVIADOS[mes], formato_subencabezado)
            hoja.write(2, columna_total_estudiantes, "TOTAL", formato_subencabezado)

            totales_dias = []
            totales_estudiantes = []
            for indice_fila, curso in enumerate(cursos, fila_datos):
                hoja.write(indice_fila, 0, curso)
                valores_dias = [len(dias[curso][mes]) for mes in meses]
                valores_estudiantes = [len(estudiantes[curso][mes]) for mes in meses]
                hoja.write_row(indice_fila, columna_dias, valores_dias, formato_numero)
                total_dias = sum(valores_dias)
                total_estudiantes = sum(valores_estudiantes)
                totales_dias.append(total_dias)
                totales_estudiantes.append(total_estudiantes)
                numero_excel = indice_fila + 1
                hoja.write_formula(
                    indice_fila, columna_total_dias,
                    f"=SUM({get_column_letter(columna_dias + 1)}{numero_excel}:{get_column_letter(columna_total_dias)}{numero_excel})",
                    formato_numero, total_dias,
                )
                hoja.write_row(indice_fila, columna_estudiantes, valores_estudiantes, formato_numero)
                hoja.write_formula(
                    indice_fila, columna_total_estudiantes,
                    f"=SUM({get_column_letter(columna_estudiantes + 1)}{numero_excel}:{get_column_letter(columna_total_estudiantes)}{numero_excel})",
                    formato_numero, total_estudiantes,
                )
                if indice_fila == fila_datos:
                    inicio_excel = fila_datos + 1
                    fin_excel = fila_total
                    hoja.write_formula(
                        indice_fila, columna_indicador,
                        f"=MAX({get_column_letter(columna_total_dias + 1)}{inicio_excel}:{get_column_letter(columna_total_dias + 1)}{fin_excel})",
                        formato_numero, max(totales_dias + [sum(len(dias[c][m]) for m in meses) for c in cursos[1:]]),
                    )

            hoja.write(fila_total, 0, "TOTAL GENERAL", formato_resumen)
            promedios_dias = []
            promedios_estudiantes = []
            for posicion, mes in enumerate(meses):
                valores = [len(dias[curso][mes]) for curso in cursos]
                promedio = sum(valores) / len(cursos)
                promedios_dias.append(promedio)
                columna = columna_dias + posicion
                hoja.write_formula(
                    fila_total, columna,
                    f"=AVERAGE({get_column_letter(columna + 1)}{fila_datos + 1}:{get_column_letter(columna + 1)}{fila_total})",
                    formato_resumen, promedio,
                )
            hoja.write_formula(
                fila_total, columna_total_dias,
                f"=AVERAGE({get_column_letter(columna_total_dias + 1)}{fila_datos + 1}:{get_column_letter(columna_total_dias + 1)}{fila_total})",
                formato_resumen, sum(totales_dias) / len(cursos),
            )
            for posicion, mes in enumerate(meses):
                valores = [len(estudiantes[curso][mes]) for curso in cursos]
                promedio = sum(valores) / len(cursos)
                promedios_estudiantes.append(promedio)
                columna = columna_estudiantes + posicion
                hoja.write_formula(
                    fila_total, columna,
                    f"=AVERAGE({get_column_letter(columna + 1)}{fila_datos + 1}:{get_column_letter(columna + 1)}{fila_total})",
                    formato_resumen, promedio,
                )
            hoja.write_formula(
                fila_total, columna_total_estudiantes,
                f"=AVERAGE({get_column_letter(columna_total_estudiantes + 1)}{fila_datos + 1}:{get_column_letter(columna_total_estudiantes + 1)}{fila_total})",
                formato_resumen, sum(totales_estudiantes) / len(cursos),
            )
            hoja.write_formula(
                fila_indicadores, columna_dias,
                f"=AVERAGE({get_column_letter(columna_dias + 1)}{fila_total + 1}:{get_column_letter(columna_total_dias)}{fila_total + 1})",
                formato_decimal, sum(promedios_dias) / cantidad_meses,
            )
            hoja.write_formula(
                fila_indicadores, columna_estudiantes,
                f"=AVERAGE({get_column_letter(columna_estudiantes + 1)}{fila_total + 1}:{get_column_letter(columna_total_estudiantes)}{fila_total + 1})",
                formato_decimal, sum(promedios_estudiantes) / cantidad_meses,
            )
            hoja.write_comment(fila_datos, columna_indicador, "Máximo del total de días por curso")
            hoja.write_comment(fila_indicadores, columna_dias, "Promedio general mensual de días")
            hoja.write_comment(fila_indicadores, columna_estudiantes, "Promedio general mensual de estudiantes")
            hoja.set_column(0, 0, 48)
            hoja.set_column(1, columna_total_estudiantes, 12)
            hoja.set_column(columna_indicador, columna_indicador, 12)
            hoja.freeze_panes(3, 1)
            hoja.autofilter(2, 0, fila_total - 1, columna_total_estudiantes)
            if progress_callback:
                progress_callback(94)
        finally:
            libro_lectura.close()
            libro_salida.close()
        ruta_nueva.replace(self.path)
        self._row_counts["Tabla Dinamica Estudiantes"] = fila_indicadores
        if progress_callback:
            progress_callback(100)
        return len(cursos), [MESES_ABREVIADOS[mes] for mes in meses]

    def crear_grafica_estudiantes(
        self, programa, periodo, progress_callback=None, resumen_actividades=None,
        diseno_cursos=None, resumen_informe=None,
    ):
        """Crea Estudiantes DG conservando fórmulas y hojas mediante copia incremental."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        libro_formulas = load_workbook(self.path, read_only=True, data_only=False)
        libro_valores = load_workbook(self.path, read_only=True, data_only=True)
        nombre_tabla = "Tabla Dinamica Estudiantes"
        if nombre_tabla not in libro_formulas.sheetnames:
            libro_formulas.close()
            libro_valores.close()
            raise ValueError(f"No existe la hoja '{nombre_tabla}'.")
        tabla_valores = libro_valores[nombre_tabla]
        filas_tabla = list(tabla_valores.iter_rows(values_only=True))
        encabezados = list(filas_tabla[2]) if len(filas_tabla) >= 3 else []
        try:
            columna_total_dias = encabezados.index("TOTAL")
        except ValueError:
            libro_formulas.close()
            libro_valores.close()
            raise ValueError("La tabla de estudiantes no contiene el bloque mensual de DÍAS.")
        columna_inicio = 1
        cantidad_meses = columna_total_dias - columna_inicio
        columna_inicio_estudiantes = columna_total_dias + 1
        fila_resumen = next(
            (
                indice for indice, fila in enumerate(filas_tabla)
                if str((fila[0] if fila else None) or "").strip().upper()
                == "TOTAL GENERAL"
            ),
            None,
        )
        if cantidad_meses <= 0 or fila_resumen is None:
            libro_formulas.close()
            libro_valores.close()
            raise ValueError("No se encontraron meses o la fila TOTAL GENERAL de estudiantes.")

        ruta_nueva = self.path.with_name("informe_con_grafica_estudiantes.xlsx")
        libro_salida = xlsxwriter.Workbook(
            ruta_nueva, {"constant_memory": True, "strings_to_urls": False}
        )
        formato_encabezado = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center", "valign": "vcenter"}
        )
        formato_subencabezado = libro_salida.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#1767A6", "align": "center", "valign": "vcenter"}
        )
        formato_numero = libro_salida.add_format({"align": "center", "valign": "vcenter"})
        formato_total = libro_salida.add_format(
            {"bold": True, "bg_color": "#E7F1FA", "font_color": "#0A3A6B", "align": "center"}
        )
        formato_resumen = libro_salida.add_format(
            {"bold": True, "bg_color": "#DDEAF6", "font_color": "#0A3A6B", "align": "center", "top": 1, "num_format": "0.0"}
        )
        hojas_excluidas = {"Estudiantes DG", "Estudiantes DG2"}
        if resumen_actividades is not None:
            hojas_excluidas.add("Tabla Dinamica Actividades")
        if resumen_informe is not None:
            hojas_excluidas.add("Resumen Informe")
        if diseno_cursos is not None:
            hojas_excluidas.add("Diseño de Cursos")
        nombres = [n for n in libro_formulas.sheetnames if n not in hojas_excluidas]
        total_copia = sum(max(libro_formulas[n].max_row or 1, 1) for n in nombres)
        copiadas = 0
        try:
            for nombre in nombres:
                origen = libro_formulas[nombre]
                origen_valores = libro_valores[nombre]
                if nombre == "Docentes DG":
                    valores_dg = [
                        (str(fila[0]), float(fila[1]))
                        for fila in origen_valores.iter_rows(min_row=2, values_only=True)
                        if fila[0] not in (None, "") and isinstance(fila[1], (int, float))
                    ]
                    agregar_docentes_dg_xlsxwriter(libro_salida, valores_dg, programa, periodo)
                    copiadas += max(origen.max_row or 1, 1)
                    continue
                destino = libro_salida.add_worksheet(nombre)
                filas_valores = origen_valores.iter_rows(values_only=True)
                for indice_fila, fila in enumerate(origen.iter_rows(values_only=True)):
                    cache = next(filas_valores, ())
                    for indice_columna, valor in enumerate(fila):
                        valor_limpio = self._excel_value(valor)
                        formato = None
                        if indice_fila == 0 and nombre != nombre_tabla:
                            formato = formato_encabezado
                        elif nombre == "Tabla Dinamica Docentes":
                            etiqueta = str(fila[0] or "").strip().upper()
                            formato = formato_resumen if etiqueta in {"PROMEDIO", "TOTAL GENERAL"} else (
                                formato_total if indice_columna == len(fila) - 1 else
                                (formato_numero if indice_columna >= 2 else None)
                            )
                        elif nombre == nombre_tabla:
                            if indice_fila == 1:
                                formato = formato_encabezado
                            elif indice_fila == 2:
                                formato = formato_subencabezado
                            elif indice_fila == fila_resumen:
                                formato = formato_resumen
                            elif indice_fila >= 3 and indice_columna >= 1:
                                formato = formato_numero
                        if isinstance(valor_limpio, str) and valor_limpio.startswith("="):
                            cache_valor = cache[indice_columna] if indice_columna < len(cache) else 0
                            destino.write_formula(
                                indice_fila, indice_columna, valor_limpio,
                                formato, 0 if cache_valor is None else cache_valor,
                            )
                        else:
                            destino.write(indice_fila, indice_columna, valor_limpio, formato)
                    copiadas += 1
                    if progress_callback and copiadas % 2_000 == 0:
                        progress_callback(min(5 + int(copiadas * 87 / total_copia), 92))
                if nombre == "Original":
                    destino.freeze_panes(1, 0)
                    if origen.max_column:
                        destino.autofilter(0, 0, max((origen.max_row or 1) - 1, 0), origen.max_column - 1)
                elif nombre == "Tabla Dinamica Docentes":
                    destino.freeze_panes(1, 2)
                    destino.set_column(0, 0, 42)
                    destino.set_column(1, 1, 34)
                    destino.set_column(2, max(origen.max_column - 1, 2), 14)
                elif nombre == nombre_tabla:
                    destino.freeze_panes(3, 1)
                    destino.set_column(0, 0, 48)
                    destino.set_column(1, max(origen.max_column - 1, 1), 12)

            agregar_estudiantes_dg_xlsxwriter(
                libro_salida, nombre_tabla, fila_resumen, columna_inicio,
                cantidad_meses, programa, periodo,
            )
            agregar_estudiantes_dg2_xlsxwriter(
                libro_salida, nombre_tabla, fila_resumen,
                columna_inicio_estudiantes, cantidad_meses, programa, periodo,
            )
            if resumen_actividades is not None:
                self._escribir_tabla_actividades(libro_salida, resumen_actividades)
            if resumen_informe is not None:
                self._escribir_resumen_informe(
                    libro_salida, resumen_informe, programa, periodo
                )
            if diseno_cursos is not None:
                self._escribir_diseno_cursos(libro_salida, *diseno_cursos)
            if progress_callback:
                progress_callback(96)
        finally:
            libro_formulas.close()
            libro_valores.close()
            libro_salida.close()
            tabla_valores = None
            origen = None
            origen_valores = None
            filas_valores = None
            gc.collect()
        ruta_nueva.replace(self.path)
        self._row_counts["Estudiantes DG"] = 0
        self._row_counts["Estudiantes DG2"] = 0
        if resumen_actividades is not None:
            self._row_counts["Tabla Dinamica Actividades"] = len(resumen_actividades)
        if resumen_informe is not None:
            self._row_counts["Resumen Informe"] = 0
        if diseno_cursos is not None:
            self._row_counts["Diseño de Cursos"] = 1
        if progress_callback:
            progress_callback(100)
        return [str(valor) for valor in encabezados[columna_inicio:columna_total_dias]]

    def crear_tabla_actividades(self, programa, periodo, progress_callback=None):
        """Cuenta registros por curso y acción para todos los roles de Original."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        libro = load_workbook(self.path, read_only=True, data_only=True)
        if "Original" not in libro.sheetnames:
            libro.close()
            raise ValueError("No existe la hoja 'Original'.")
        hoja = libro["Original"]
        filas = hoja.iter_rows(values_only=True)
        encabezados = next(filas, None)
        columnas = {
            str(valor or "").strip().casefold(): indice
            for indice, valor in enumerate(encabezados or ())
        }
        faltantes = [nombre for nombre in ("curso", "accion", "rol") if nombre not in columnas]
        if faltantes:
            libro.close()
            raise ValueError(
                "La hoja Original no contiene las columnas requeridas: "
                + ", ".join(faltantes)
            )
        resumen = defaultdict(lambda: defaultdict(int))
        total = max((hoja.max_row or 1) - 1, 1)
        try:
            for indice, fila in enumerate(filas, 1):
                curso = str(fila[columnas["curso"]] or "").strip()
                accion = str(fila[columnas["accion"]] or "").strip().casefold()
                if curso and accion:
                    resumen[curso][accion] += 1
                if progress_callback and indice % 5_000 == 0:
                    progress_callback(min(5 + int(indice * 40 / total), 45))
        finally:
            libro.close()
        if not resumen:
            raise ValueError("No se encontraron cursos con acciones en la hoja Original.")
        self.crear_grafica_estudiantes(
            programa, periodo,
            (lambda valor: progress_callback(45 + int(valor * 0.54)))
            if progress_callback else None,
            resumen,
        )
        if progress_callback:
            progress_callback(100)
        acciones = sorted({accion for valores in resumen.values() for accion in valores})
        return len(resumen), acciones

    @staticmethod
    def _escribir_tabla_actividades(libro, resumen):
        """Escribe la tabla equivalente a la dinámica del informe de referencia."""
        hoja = libro.add_worksheet("Tabla Dinamica Actividades")
        acciones = sorted({accion for valores in resumen.values() for accion in valores})
        azul = libro.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B", "align": "center"}
        )
        subazul = libro.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#1767A6", "align": "center"}
        )
        numero = libro.add_format({"align": "center", "num_format": "0"})
        total_fmt = libro.add_format(
            {"bold": True, "font_color": "#082B55", "bg_color": "#DDEAF6", "align": "center", "top": 1}
        )
        hoja.write_row(0, 0, ["rol", "All"], azul)
        hoja.write_row(2, 0, ["Recuento de accion", "Etiquetas de columna"], azul)
        hoja.write_row(3, 0, ["Etiquetas de fila", *acciones, "Total general"], subazul)
        totales = {accion: 0 for accion in acciones}
        for fila, curso in enumerate(sorted(resumen, key=str.casefold), 4):
            hoja.write(fila, 0, curso)
            total_curso = 0
            for columna, accion in enumerate(acciones, 1):
                valor = resumen[curso].get(accion, 0)
                if valor:
                    hoja.write_number(fila, columna, valor, numero)
                total_curso += valor
                totales[accion] += valor
            hoja.write_number(fila, len(acciones) + 1, total_curso, total_fmt)
        fila_total = 4 + len(resumen)
        hoja.write(fila_total, 0, "Total general", total_fmt)
        for columna, accion in enumerate(acciones, 1):
            hoja.write_number(fila_total, columna, totales[accion], total_fmt)
        hoja.write_number(fila_total, len(acciones) + 1, sum(totales.values()), total_fmt)
        hoja.set_column(0, 0, 48)
        hoja.set_column(1, len(acciones) + 1, 15)
        hoja.freeze_panes(4, 1)
        hoja.autofilter(3, 0, fila_total - 1, len(acciones) + 1)

    def crear_diseno_cursos(self, programa, periodo, progress_callback=None):
        """Resume los cursos unificados y los cursos con o sin actividad."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        libro = load_workbook(self.path, read_only=True, data_only=True)
        nombre = "Tabla Dinamica Actividades"
        if nombre not in libro.sheetnames:
            libro.close()
            raise ValueError(f"No existe la hoja '{nombre}'.")
        hoja = libro[nombre]
        encabezados = [str(hoja.cell(4, columna).value or "").strip() for columna in range(1, hoja.max_column + 1)]
        acciones = encabezados[1:-1]
        resumen = defaultdict(lambda: defaultdict(int))
        for fila in range(5, hoja.max_row + 1):
            curso = str(hoja.cell(fila, 1).value or "").strip()
            if not curso or curso.casefold() == "total general":
                continue
            for columna, accion in enumerate(acciones, 2):
                valor = hoja.cell(fila, columna).value
                if isinstance(valor, (int, float)) and valor:
                    resumen[curso][accion.casefold()] = int(valor)
        resumen_informe = self._calcular_resumen_informe(
            libro["Original"],
            (lambda valor: progress_callback(valor)) if progress_callback else None,
        )
        libro.close()
        if not resumen:
            raise ValueError("La tabla de actividades no contiene cursos para evaluar.")

        patron_unificado = re.compile(r"_\d+[A-Z]_\d+[A-Z]$", re.IGNORECASE)
        unificados = {
            curso for curso in resumen
            if patron_unificado.search(curso)
            and "PRACTICA EMPRESARIAL" not in curso.upper()
        }
        cursos_evaluados = [curso for curso in resumen if curso not in unificados]
        con_contenido = sum(bool(sum(resumen[curso].values())) for curso in cursos_evaluados)
        sin_contenido = len(cursos_evaluados) - con_contenido
        indicadores = (len(unificados), sin_contenido, con_contenido)
        self.crear_grafica_estudiantes(
            programa, periodo, progress_callback, resumen, indicadores,
            resumen_informe,
        )
        return indicadores

    @staticmethod
    def _calcular_resumen_informe(hoja_original, progress_callback=None):
        """Consolida los indicadores que alimentaran Resumen Informe."""
        filas = hoja_original.iter_rows(values_only=True)
        encabezados = next(filas, None)
        columnas = {
            str(valor or "").strip().casefold(): indice
            for indice, valor in enumerate(encabezados or ())
        }
        requeridas = ("rol", "curso", "mes", "dia", "idusuario")
        faltantes = [nombre for nombre in requeridas if nombre not in columnas]
        if faltantes:
            raise ValueError(
                "Original no contiene las columnas para Resumen Informe: "
                + ", ".join(faltantes)
            )

        eventos_mes = defaultdict(Counter)
        usuarios_mes = defaultdict(set)
        usuarios_rol = defaultdict(set)
        dias_usuario = defaultdict(set)
        eventos_curso = Counter()
        usuarios_curso = defaultdict(set)
        dias_curso = defaultdict(set)
        dias_docente_curso = defaultdict(set)
        total_eventos = 0

        total_filas = max((hoja_original.max_row or 1) - 1, 1)
        for numero_fila, fila in enumerate(filas, 1):
            total_eventos += 1
            rol = str(fila[columnas["rol"]] or "").strip().casefold()
            curso = str(fila[columnas["curso"]] or "").strip()
            usuario = fila[columnas["idusuario"]]
            if usuario in (None, ""):
                continue
            usuario = str(usuario).strip()
            nombre_docente = (
                str(fila[columnas["usuario"]] or "").strip()
                if "usuario" in columnas else usuario
            )
            try:
                mes = int(float(fila[columnas["mes"]]))
                dia = int(float(fila[columnas["dia"]]))
            except (TypeError, ValueError):
                continue
            if not 1 <= mes <= 12 or not 1 <= dia <= 31:
                continue
            fecha = (mes, dia)
            usuarios_rol[rol].add(usuario)
            dias_usuario[(rol, usuario)].add(fecha)
            eventos_mes[mes][rol] += 1
            if rol == "student":
                usuarios_mes[mes].add(usuario)
                if curso:
                    eventos_curso[curso] += 1
                    usuarios_curso[curso].add(usuario)
                    dias_curso[curso].add(fecha)
            elif rol == "editingteacher" and curso:
                dias_docente_curso[(curso, nombre_docente or usuario)].add(fecha)
            if progress_callback and numero_fila % 5_000 == 0:
                progress_callback(min(1 + int(numero_fila * 3 / total_filas), 4))

        def estadistica_dias(rol):
            valores = [
                len(dias) for (rol_fila, _), dias in dias_usuario.items()
                if rol_fila == rol
            ]
            if not valores:
                return 0.0, 0.0
            return round(sum(valores) / len(valores), 1), float(median(valores))

        promedio_estudiantes, mediana_estudiantes = estadistica_dias("student")
        promedio_docentes, mediana_docentes = estadistica_dias("editingteacher")
        meses = sorted(eventos_mes)
        cursos = [
            {
                "curso": curso,
                "eventos": cantidad,
                "estudiantes": len(usuarios_curso[curso]),
                "dias": len(dias_curso[curso]),
            }
            for curso, cantidad in eventos_curso.most_common(6)
        ]
        docentes = [
            {"curso": curso, "docente": docente, "dias": len(dias)}
            for (curso, docente), dias in sorted(
                dias_docente_curso.items(),
                key=lambda elemento: (-len(elemento[1]), elemento[0][0].casefold()),
            )[:6]
        ]
        return {
            "total_eventos": total_eventos,
            "usuarios_unicos": len(set().union(*usuarios_rol.values())),
            "estudiantes": len(usuarios_rol["student"]),
            "docentes": len(usuarios_rol["editingteacher"]),
            "eventos_estudiantes": sum(eventos_mes[m]["student"] for m in meses),
            "eventos_docentes": sum(eventos_mes[m]["editingteacher"] for m in meses),
            "promedio_dias_estudiantes": promedio_estudiantes,
            "mediana_dias_estudiantes": mediana_estudiantes,
            "promedio_dias_docentes": promedio_docentes,
            "mediana_dias_docentes": mediana_docentes,
            "meses": [
                {
                    "mes": MESES_ABREVIADOS[mes],
                    "eventos_estudiantes": eventos_mes[mes]["student"],
                    "estudiantes_activos": len(usuarios_mes[mes]),
                    "eventos_docentes": eventos_mes[mes]["editingteacher"],
                }
                for mes in meses
            ],
            "cursos": cursos,
            "docentes_destacados": docentes,
        }

    @staticmethod
    def _escribir_resumen_informe(libro, datos, programa, periodo):
        """Crea la hoja ejecutiva que sirve de fuente para el informe Word."""
        hoja = libro.add_worksheet("Resumen Informe")
        hoja.set_tab_color("#D6A419")
        hoja.hide_gridlines(2)
        hoja.set_landscape()
        hoja.fit_to_pages(1, 2)
        hoja.freeze_panes(4, 0)

        titulo = libro.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#082B55",
            "align": "center", "valign": "vcenter", "font_size": 18,
        })
        seccion = libro.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#1767A6",
            "align": "left", "valign": "vcenter", "font_size": 12,
        })
        encabezado = libro.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#0A3A6B",
            "align": "center", "valign": "vcenter", "border": 1,
        })
        texto = libro.add_format({
            "font_color": "#23364D", "bg_color": "#FFFFFF", "border": 1,
            "border_color": "#C7D4E3", "valign": "vcenter",
        })
        numero = libro.add_format({
            "font_color": "#23364D", "bg_color": "#F4F7FB", "border": 1,
            "border_color": "#C7D4E3", "align": "center", "num_format": "#,##0",
        })
        decimal = libro.add_format({
            "font_color": "#23364D", "bg_color": "#F4F7FB", "border": 1,
            "border_color": "#C7D4E3", "align": "center", "num_format": "0.0",
        })

        hoja.merge_range("A1:J2", f"Resumen del informe Open LMS - {programa} {periodo}", titulo)
        hoja.merge_range("A4:C4", "Indicadores principales", seccion)
        hoja.write_row("A5", ["Indicador", "Resultado", "Lectura"], encabezado)
        filas_indicadores = [
            ("Eventos totales", datos["total_eventos"], "Interacciones registradas"),
            ("Usuarios \u00fanicos", datos["usuarios_unicos"], f'{datos["estudiantes"]} estudiantes y {datos["docentes"]} docentes'),
            ("Actividad estudiantil", datos["eventos_estudiantes"], "Eventos de estudiantes"),
            ("Actividad docente", datos["eventos_docentes"], "Eventos de docentes"),
            ("D\u00edas activos por estudiante", datos["promedio_dias_estudiantes"], f'Mediana: {datos["mediana_dias_estudiantes"]:g} d\u00edas'),
            ("D\u00edas activos por docente", datos["promedio_dias_docentes"], f'Mediana: {datos["mediana_dias_docentes"]:g} d\u00edas'),
        ]
        for fila, (indicador, valor, lectura) in enumerate(filas_indicadores, 5):
            hoja.write(fila, 0, indicador, texto)
            hoja.write_number(fila, 1, valor, decimal if isinstance(valor, float) and not valor.is_integer() else numero)
            hoja.write(fila, 2, lectura, texto)

        fila_meses = 13
        hoja.merge_range(fila_meses, 0, fila_meses, 3, "Actividad mensual", seccion)
        hoja.write_row(fila_meses + 1, 0, ["Mes", "Eventos estudiantes", "Estudiantes activos", "Eventos docentes"], encabezado)
        for indice, mes in enumerate(datos["meses"], fila_meses + 2):
            hoja.write(indice, 0, mes["mes"], texto)
            hoja.write_row(indice, 1, [mes["eventos_estudiantes"], mes["estudiantes_activos"], mes["eventos_docentes"]], numero)

        grafica_mensual = libro.add_chart({"type": "line"})
        ultima_fila_mes = fila_meses + 1 + len(datos["meses"])
        for columna, nombre, color in ((1, "Estudiantes", "#1767A6"), (3, "Docentes", "#D6A419")):
            grafica_mensual.add_series({
                "name": nombre,
                "categories": ["Resumen Informe", fila_meses + 2, 0, ultima_fila_mes, 0],
                "values": ["Resumen Informe", fila_meses + 2, columna, ultima_fila_mes, columna],
                "line": {"color": color, "width": 2.5},
                "marker": {"type": "circle", "size": 7, "fill": {"color": color}},
            })
        grafica_mensual.set_title({"name": "Eventos mensuales por tipo de usuario"})
        grafica_mensual.set_y_axis({"name": "Eventos registrados", "major_gridlines": {"visible": True}})
        grafica_mensual.set_legend({"position": "top"})
        grafica_mensual.set_size({"width": 700, "height": 330})
        hoja.insert_chart("F4", grafica_mensual)

        fila_cursos = 23
        hoja.merge_range(fila_cursos, 0, fila_cursos, 3, "Cursos destacados", seccion)
        hoja.write_row(fila_cursos + 1, 0, ["Curso", "Eventos", "Estudiantes \u00fanicos", "D\u00edas activos"], encabezado)
        for indice, curso in enumerate(datos["cursos"], fila_cursos + 2):
            hoja.write(indice, 0, curso["curso"], texto)
            hoja.write_row(indice, 1, [curso["eventos"], curso["estudiantes"], curso["dias"]], numero)

        grafica_cursos = libro.add_chart({"type": "bar"})
        ultima_fila_curso = fila_cursos + 1 + len(datos["cursos"])
        grafica_cursos.add_series({
            "name": "Eventos estudiantiles",
            "categories": ["Resumen Informe", fila_cursos + 2, 0, ultima_fila_curso, 0],
            "values": ["Resumen Informe", fila_cursos + 2, 1, ultima_fila_curso, 1],
            "fill": {"color": "#5B9BD5"}, "border": {"none": True},
            "data_labels": {"value": True, "num_format": "#,##0"},
        })
        grafica_cursos.set_title({"name": "Cursos con mayor actividad estudiantil"})
        grafica_cursos.set_x_axis({"name": "Eventos estudiantiles"})
        grafica_cursos.set_legend({"none": True})
        grafica_cursos.set_size({"width": 700, "height": 360})
        hoja.insert_chart("F22", grafica_cursos)

        fila_docentes = 34
        hoja.merge_range(fila_docentes, 0, fila_docentes, 2, "Continuidad docente", seccion)
        hoja.write_row(fila_docentes + 1, 0, ["Curso", "Docente", "D\u00edas activos"], encabezado)
        for indice, docente in enumerate(datos["docentes_destacados"], fila_docentes + 2):
            hoja.write(indice, 0, docente["curso"], texto)
            hoja.write(indice, 1, docente["docente"], texto)
            hoja.write_number(indice, 2, docente["dias"], numero)

        hoja.set_column("A:A", 46)
        hoja.set_column("B:B", 24)
        hoja.set_column("C:D", 19)
        hoja.set_column("E:E", 3)
        hoja.set_column("F:J", 14)
        hoja.set_row(0, 28)

    @staticmethod
    def _escribir_diseno_cursos(libro, unificados, sin_contenido, con_contenido):
        """Crea el resumen final y su gráfico circular institucional."""
        hoja = libro.add_worksheet("Diseño de Cursos")
        titulo = libro.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#082B55", "align": "center", "font_size": 14}
        )
        encabezado = libro.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#1767A6", "align": "center"}
        )
        valor = libro.add_format(
            {"bold": True, "font_color": "#082B55", "bg_color": "#F4F7FB", "align": "center", "font_size": 12, "border": 1, "border_color": "#C7D4E3"}
        )
        hoja.merge_range("C4:F4", "Diseño cursos", titulo)
        hoja.write_row("C5", ["Unificados", "Sin contenido", "Con contenido", "Total"], encabezado)
        hoja.write_row("C6", [unificados, sin_contenido, con_contenido, sin_contenido + con_contenido], valor)
        hoja.set_column("A:B", 3)
        hoja.set_column("C:F", 18)
        hoja.set_row(3, 26)
        grafica = libro.add_chart({"type": "pie"})
        grafica.add_series(
            {
                "name": "Diseño de cursos",
                "categories": ["Diseño de Cursos", 4, 3, 4, 4],
                "values": ["Diseño de Cursos", 5, 3, 5, 4],
                "points": [
                    {"fill": {"color": "#D6A419"}},
                    {"fill": {"color": "#1767A6"}},
                ],
                "data_labels": {
                    "percentage": True, "category": True,
                    "leader_lines": True,
                    "font": {"color": "#082B55", "bold": True},
                },
            }
        )
        grafica.set_title(
            {"name": "Diseño de Cursos", "name_font": {"color": "#082B55", "bold": True, "size": 16}}
        )
        grafica.set_legend({"position": "bottom"})
        grafica.set_chartarea({"fill": {"color": "#F4F7FB"}, "border": {"color": "#C7D4E3"}})
        grafica.set_size({"width": 720, "height": 430})
        hoja.insert_chart("C8", grafica)

    def preview_sheet(self, sheet_name, limit=200):
        libro_excel = load_workbook(self.path, read_only=True, data_only=True)
        try:
            hoja = libro_excel[sheet_name]
            total_filas = self._row_counts.get(sheet_name)
            if total_filas is None:
                total_filas = max((hoja.max_row or 1) - 1, 0)
            filas = []
            for indice, fila in enumerate(hoja.iter_rows(values_only=True)):
                if indice > limit:
                    break
                filas.append(fila)
        finally:
            libro_excel.close()
        if not filas:
            return [], [], total_filas
        encabezados = [str(valor or "") for valor in filas[0]]
        return encabezados, filas[1:], total_filas

    def save_as(self, destination):
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        shutil.copy2(self.path, destination)

    def finalize_as(self, destination):
        """Guarda el libro definitivo y elimina el temporal solo al finalizar bien."""
        if not self.exists:
            raise FileNotFoundError("Todavía no existe un Excel de trabajo.")
        origen = self.path.resolve()
        destino = Path(destination).resolve()
        if origen == destino:
            return
        destino.parent.mkdir(parents=True, exist_ok=True)
        descriptor, nombre_temporal = tempfile.mkstemp(
            prefix=f".{destino.stem}_", suffix=".xlsx", dir=destino.parent
        )
        import os
        os.close(descriptor)
        copia_temporal = Path(nombre_temporal)
        try:
            shutil.copy2(origen, copia_temporal)
            copia_temporal.replace(destino)
            origen.unlink()
        finally:
            copia_temporal.unlink(missing_ok=True)
        self.path = destino


def suggested_filename(period, program):
    periodo_seguro = re.sub(r"[^\w-]+", "_", period.strip(), flags=re.UNICODE)
    programa_seguro = re.sub(r"[^\w-]+", "_", program.strip(), flags=re.UNICODE).strip("_")
    return f"Informe_{periodo_seguro}_{programa_seguro}.xlsx"
