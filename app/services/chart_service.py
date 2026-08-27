"""Gráficas nativas que se incorporan al libro Excel de trabajo."""

from numbers import Number

from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


HOJA_ORIGEN_DOCENTES = "Tabla Dinamica Docentes"
HOJA_GRAFICA_DOCENTES = "Docentes DG"
HOJA_GRAFICA_ESTUDIANTES = "Estudiantes DG"
HOJA_GRAFICA_ESTUDIANTES_2 = "Estudiantes DG2"
MESES_VALIDOS = {
    "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGO", "SEP", "OCT", "NOV", "DIC",
}


def crear_docentes_dg(libro, programa, periodo):
    """Crea la tabla y gráfica de promedios mensuales desde la tabla docente."""
    if HOJA_ORIGEN_DOCENTES not in libro.sheetnames:
        raise ValueError(
            f"No existe la hoja '{HOJA_ORIGEN_DOCENTES}'. Créala antes de generar la gráfica."
        )
    origen = libro[HOJA_ORIGEN_DOCENTES]
    encabezados = [str(celda.value or "").strip() for celda in origen[1]]
    columnas_mensuales = [
        (indice, nombre.upper())
        for indice, nombre in enumerate(encabezados, 1)
        if nombre.upper() in MESES_VALIDOS
    ]
    if not columnas_mensuales:
        raise ValueError(
            f"La hoja '{HOJA_ORIGEN_DOCENTES}' no contiene columnas mensuales."
        )

    promedios = []
    for indice_columna, mes in columnas_mensuales:
        valores = []
        for fila in range(2, origen.max_row + 1):
            etiqueta = str(origen.cell(fila, 1).value or "").strip().upper()
            if etiqueta in {"PROMEDIO", "TOTAL GENERAL"}:
                continue
            valor = origen.cell(fila, indice_columna).value
            if isinstance(valor, Number) and not isinstance(valor, bool):
                valores.append(float(valor))
        if valores:
            promedios.append((mes, round(sum(valores) / len(valores), 1)))
    if not promedios:
        raise ValueError(
            f"Las columnas mensuales de '{HOJA_ORIGEN_DOCENTES}' no contienen valores numéricos."
        )

    if HOJA_GRAFICA_DOCENTES in libro.sheetnames:
        libro.remove(libro[HOJA_GRAFICA_DOCENTES])
    hoja = libro.create_sheet(HOJA_GRAFICA_DOCENTES)
    hoja.append(["MES", "PROMEDIO"])
    for mes, promedio in promedios:
        hoja.append([mes, promedio])

    relleno = PatternFill("solid", fgColor="0A3A6B")
    for celda in hoja[1]:
        celda.fill = relleno
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center")
    for fila in range(2, hoja.max_row + 1):
        hoja.cell(fila, 1).alignment = Alignment(horizontal="center")
        hoja.cell(fila, 2).alignment = Alignment(horizontal="center")
        hoja.cell(fila, 2).number_format = "0.0"
    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 18
    hoja.freeze_panes = "A2"

    grafica = LineChart()
    grafica.title = (
        "Días al mes de uso del Campus Virtual por parte de los docentes\n"
        f"de la facultad de {programa} {periodo}"
    )
    grafica.x_axis.title = "Mes"
    grafica.y_axis.title = "Promedio de días de uso"
    grafica.height = 10
    grafica.width = 20
    grafica.style = 13
    datos = Reference(hoja, min_col=2, min_row=1, max_row=hoja.max_row)
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=hoja.max_row)
    grafica.add_data(datos, titles_from_data=True)
    grafica.set_categories(categorias)
    grafica.legend = None
    serie = grafica.series[0]
    serie.marker.symbol = "circle"
    serie.marker.size = 7
    serie.graphicalProperties.line.solidFill = "0A4D91"
    serie.graphicalProperties.line.width = 28575
    hoja.add_chart(grafica, "D2")
    return promedios


def agregar_docentes_dg_xlsxwriter(libro, promedios, programa, periodo):
    """Escribe Docentes DG sin cargar en memoria las demás hojas del libro."""
    hoja = libro.add_worksheet(HOJA_GRAFICA_DOCENTES)
    encabezado = libro.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#0A3A6B",
            "align": "center",
            "valign": "vcenter",
        }
    )
    promedio_formato = libro.add_format(
        {"align": "center", "valign": "vcenter", "num_format": "0.0"}
    )
    hoja.write_row(0, 0, ["MES", "PROMEDIO"], encabezado)
    for indice, (mes, promedio) in enumerate(promedios, 1):
        hoja.write(indice, 0, mes, promedio_formato)
        hoja.write_number(indice, 1, promedio, promedio_formato)
    hoja.set_column("A:A", 14)
    hoja.set_column("B:B", 18)
    hoja.freeze_panes(1, 0)
    hoja.set_tab_color("#1767A6")

    grafica = libro.add_chart({"type": "line"})
    grafica.add_series(
        {
            "name": "PROMEDIO",
            "categories": [HOJA_GRAFICA_DOCENTES, 1, 0, len(promedios), 0],
            "values": [HOJA_GRAFICA_DOCENTES, 1, 1, len(promedios), 1],
            "line": {"color": "#1767A6", "width": 2.75},
            "marker": {
                "type": "circle",
                "size": 10,
                "border": {"color": "#FFFFFF", "width": 1},
                "fill": {"color": "#1767A6"},
            },
            "data_labels": {
                "value": True,
                "num_format": "0.0",
                "position": "center",
                "font": {"color": "#FFFFFF", "bold": True, "size": 9},
                "fill": {"color": "#1767A6"},
                "border": {"color": "#1767A6"},
            },
        }
    )
    grafica.set_title(
        {
            "name": "Días al mes de uso del Campus Virtual por parte de los docentes\n"
            f"de la facultad de {programa} {periodo}",
            "name_font": {"color": "#082B55", "bold": True, "size": 16},
        }
    )
    grafica.set_x_axis(
        {
            "name": "Mes",
            "label_position": "low",
            "line": {"color": "#555555", "width": 1.5},
            "major_tick_mark": "none",
        }
    )
    grafica.set_y_axis(
        {
            "name": "Promedio de días de uso",
            "num_format": "0.0",
            "label_position": "none",
            "line": {"none": True},
            "major_gridlines": {"visible": True, "line": {"color": "#D9E2EC"}},
            "major_tick_mark": "none",
        }
    )
    grafica.set_legend({"none": True})
    grafica.set_chartarea(
        {"fill": {"color": "#F4F7FB"}, "border": {"color": "#C7D4E3"}}
    )
    grafica.set_plotarea(
        {"fill": {"color": "#FFFFFF"}, "border": {"none": True}}
    )
    grafica.set_size({"width": 900, "height": 540})
    hoja.insert_chart("B2", grafica)


def agregar_estudiantes_dg_xlsxwriter(
    libro, hoja_origen, fila_resumen, columna_inicio, cantidad_meses,
    programa, periodo,
):
    """Agrega la gráfica mensual de días de uso de estudiantes."""
    _agregar_grafica_estudiantes_xlsxwriter(
        libro, HOJA_GRAFICA_ESTUDIANTES, hoja_origen, fila_resumen,
        columna_inicio, cantidad_meses,
        "Días del mes de uso del Campus Virtual por parte de los estudiantes "
        f"de la facultad de {programa} {periodo}",
        "Promedio de días de uso", "#1767A6", "0.0",
    )


def agregar_estudiantes_dg2_xlsxwriter(
    libro, hoja_origen, fila_resumen, columna_inicio, cantidad_meses,
    programa, periodo,
):
    """Agrega la gráfica mensual del promedio de estudiantes distintos."""
    _agregar_grafica_estudiantes_xlsxwriter(
        libro, HOJA_GRAFICA_ESTUDIANTES_2, hoja_origen, fila_resumen,
        columna_inicio, cantidad_meses,
        "Promedio de estudiantes que usaron el Campus Virtual de la facultad "
        f"de {programa} {periodo}",
        "Promedio de estudiantes", "#D6A419", "0.0",
    )


def _agregar_grafica_estudiantes_xlsxwriter(
    libro, nombre_hoja, hoja_origen, fila_resumen, columna_inicio,
    cantidad_meses, titulo, nombre_serie, color_serie, formato_numero,
):
    """Construye una gráfica de línea con una apariencia institucional común."""
    hoja = libro.add_worksheet(nombre_hoja)
    hoja.set_tab_color(color_serie)
    grafica = libro.add_chart({"type": "line"})
    grafica.add_series(
        {
            "name": nombre_serie,
            "categories": [hoja_origen, 2, columna_inicio, 2, columna_inicio + cantidad_meses - 1],
            "values": [
                hoja_origen, fila_resumen, columna_inicio,
                fila_resumen, columna_inicio + cantidad_meses - 1,
            ],
            "line": {"color": color_serie, "width": 2.75},
            "marker": {
                "type": "circle", "size": 10,
                "border": {"color": "#FFFFFF", "width": 1},
                "fill": {"color": color_serie},
            },
            "data_labels": {
                "value": True, "num_format": formato_numero, "position": "center",
                "font": {"color": "#FFFFFF", "bold": True, "size": 9},
                "fill": {"color": color_serie},
                "border": {"color": color_serie},
            },
        }
    )
    grafica.set_title({"name": titulo, "name_font": {"color": "#082B55", "bold": True, "size": 16}})
    grafica.set_x_axis(
        {"major_tick_mark": "none", "line": {"color": "#555555", "width": 1.5}}
    )
    grafica.set_y_axis(
        {
            "name": nombre_serie, "num_format": formato_numero, "label_position": "none",
            "line": {"none": True}, "major_tick_mark": "none",
            "major_gridlines": {"visible": True, "line": {"color": "#D9E2EC"}},
        }
    )
    grafica.set_legend({"none": True})
    grafica.set_chartarea({"fill": {"color": "#F4F7FB"}, "border": {"color": "#C7D4E3"}})
    grafica.set_plotarea({"fill": {"color": "#FFFFFF"}, "border": {"none": True}})
    grafica.set_size({"width": 900, "height": 540})
    hoja.insert_chart("B2", grafica)
