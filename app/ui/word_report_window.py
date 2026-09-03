"""Explorador interactivo para convertir Excel terminados en informes Word."""

from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import QObject, QSize, QThread, Qt, QUrl, Signal, Slot
from PySide6.QtGui import QColor, QDesktopServices, QFont, QIcon, QPainter, QPixmap
from PySide6.QtWidgets import (
    QComboBox, QFileDialog, QFrame, QGridLayout, QHBoxLayout, QLabel,
    QLineEdit, QListWidget, QListWidgetItem, QProgressBar, QPushButton,
    QSizePolicy, QVBoxLayout, QWidget,
)

from app.services.report_option_service import list_report_options
from app.services.report_path_service import build_pdf_path, build_word_path
from app.services.pdf_report_service import convert_word_to_pdf
from app.services.process_history_service import list_completed_processes
from app.services.word_report_service import generate_word_report
from app.ui.modal_dialogs import ask_confirmation, show_error, show_info
from app.ui.theme import EXCEL_MODULE_STYLESHEET


WORD_STYLE = EXCEL_MODULE_STYLESHEET + """
QFrame#wordHero { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #123F91, stop:0.55 #0759B6, stop:1 #261478); border: none; border-radius: 16px; }
QLabel#wordHeroImage { background: transparent; border: none; }
QLabel#wordEyebrow { color: #78E3F8; font-size: 9px; font-weight: 900; letter-spacing: 2px; }
QLabel#wordTitle { color: #FFFFFF; font-size: 26px; font-weight: 800; }
QLabel#wordSubtitle { color: #E5F1FF; font-size: 11px; }
QLabel#wordStepNumber { color: #0D4380; background-color: #8EE8F7; border-radius: 10px; font-size: 9px; font-weight: 900; }
QLabel#wordStepText { color: #FFFFFF; font-size: 9px; font-weight: 800; }
QLabel#wordStepArrow { color: #78CDEB; font-size: 11px; }
QFrame#wordPanel { background-color: #FFFFFF; border: 1px solid #D9E2EC; border-radius: 14px; }
QLabel#wordSection { color: #102A49; font-size: 16px; font-weight: 800; }
QLabel#wordPanelHint { color: #718096; font-size: 10px; }
QLabel#wordMuted { color: #718096; font-size: 10px; }
QLineEdit#reportSearch, QComboBox#reportFilter {
    background-color: #F8FAFC; color: #173653; border: 1px solid #D3DDE8;
    border-radius: 9px; padding: 9px 12px; min-height: 22px;
}
QLineEdit#reportSearch:focus, QComboBox#reportFilter:focus { border: 2px solid #0B67D1; }
QComboBox#reportFilter { color: #0B67D1; font-weight: 800; background-color: #F4F8FF; }
QListWidget#reportList { background: #F5F8FB; border: 1px solid #DFE7EF; border-radius: 10px; padding: 6px; outline: 0; }
QListWidget#reportList::item { color: #173653; background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 9px; padding: 12px; margin: 3px; }
QListWidget#reportList::item:hover { border-color: #83B8EA; background: #F3F8FE; }
QListWidget#reportList::item:selected { color: #073A6F; border: 2px solid #0B67D1; background: #EAF3FC; }
QLabel#countBadge { color: #075EAE; background: #E5F2FF; border-radius: 10px; padding: 4px 9px; font-weight: 800; }
QLabel#selectionTitle { color: #0B315A; font-size: 18px; font-weight: 800; }
QLabel#selectionBadge { color: #137346; background-color: #E3F4EA; border-radius: 10px; padding: 5px 10px; font-size: 9px; font-weight: 900; }
QLabel#detailLabel { color: #64748B; font-size: 10px; font-weight: 700; }
QLabel#detailValue { color: #183455; font-size: 11px; background: #F5F8FB; border: 1px solid #E7EDF3; border-radius: 8px; padding: 9px; }
QFrame#wordActionArea { background-color: #F5F9FD; border: 1px solid #DDE8F2; border-radius: 10px; }
QPushButton#wordPrimary { background: #0B67D1; color: white; border: none; border-radius: 9px; padding: 12px 18px; font-size: 12px; font-weight: 900; }
QPushButton#wordPrimary:hover { background: #0959B7; }
QPushButton#wordPrimary:disabled { background: #D8E0E9; color: #929EAC; }
QPushButton#wordAction { background: #FFFFFF; color: #0B5DAC; border: 1px solid #B7CCE1; border-radius: 8px; padding: 9px 14px; font-weight: 800; }
QPushButton#wordAction:hover { background: #EDF6FF; }
QPushButton#wordSelectAction { background: #176CE0; color: #FFFFFF; border: none; border-radius: 8px; padding: 9px 14px; font-weight: 900; }
QPushButton#wordSelectAction:hover { background: #0D5FCB; }
"""


class WordGenerationTask(QObject):
    finished = Signal(object)
    failed = Signal(str)

    def __init__(self, operation):
        super().__init__()
        self.operation = operation

    @Slot()
    def run(self):
        try:
            result = self.operation()
        except BaseException as error:  # noqa: BLE001 - debe mostrarse, no tumbar la app
            self.failed.emit(f"{type(error).__name__}: {error}")
        else:
            self.finished.emit(result)


class WordReportWindow(QWidget):
    back_requested = Signal()

    def __init__(self, user):
        super().__init__()
        self.user = user
        self.setObjectName("excelProcessPage")
        self.setStyleSheet(WORD_STYLE)
        self.reports = []
        self.excel_path = None
        self.generated_path = None
        self.generated_pdf_path = None
        self.current_program = ""
        self.current_period = ""
        self._thread = None
        self._worker = None
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(30, 20, 30, 20)
        root.setSpacing(13)
        back = QPushButton("<-  Volver al dashboard", objectName="excelBackButton")
        back.clicked.connect(self.back_requested.emit)
        root.addWidget(back, alignment=Qt.AlignLeft)

        hero = QFrame(objectName="wordHero")
        hero_box = QHBoxLayout(hero)
        hero_box.setContentsMargins(25, 14, 25, 14)
        hero_box.setSpacing(12)
        hero_copy = QVBoxLayout()
        hero_copy.setSpacing(2)
        eyebrow = QLabel("GENERADOR INSTITUCIONAL  /  EXCEL A WORD", objectName="wordEyebrow")
        title = QLabel("Construye el informe final", objectName="wordTitle")
        for label in (eyebrow, title):
            label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        hero_copy.addWidget(eyebrow)
        hero_copy.addWidget(title)
        subtitle = QLabel("Busca un Excel terminado, revisa sus datos y genera el documento con graficos y plantilla institucional.", objectName="wordSubtitle")
        subtitle.setWordWrap(True)
        subtitle.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        hero_copy.addWidget(subtitle)
        steps = QHBoxLayout()
        steps.setSpacing(7)
        for index, step_text in enumerate(("Seleccionar", "Revisar", "Generar"), 1):
            number = QLabel(str(index), objectName="wordStepNumber")
            number.setAlignment(Qt.AlignCenter)
            number.setFixedSize(20, 20)
            steps.addWidget(number)
            steps.addWidget(QLabel(step_text, objectName="wordStepText"))
            if index < 3:
                steps.addWidget(QLabel(">", objectName="wordStepArrow"))
        steps.addStretch()
        hero_copy.addLayout(steps)
        hero_box.addLayout(hero_copy, 4)
        hero_image = QLabel(objectName="wordHeroImage")
        hero_image.setAlignment(Qt.AlignCenter)
        hero_image.setPixmap(
            QPixmap(str(Path(__file__).parent / "assets" / "excel-to-word-hero.png")).scaled(
                330, 135, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
        )
        hero_image.setFixedSize(340, 135)
        hero_box.addWidget(hero_image, 3)
        root.addWidget(hero)

        content = QHBoxLayout()
        content.setSpacing(13)
        root.addLayout(content, 1)
        explorer = QFrame(objectName="wordPanel")
        left = QVBoxLayout(explorer)
        left.setContentsMargins(18, 16, 18, 16)
        heading = QHBoxLayout()
        heading.addWidget(QLabel("Excel terminados", objectName="wordSection"))
        self.count_badge = QLabel("0 disponibles", objectName="countBadge")
        heading.addStretch()
        heading.addWidget(self.count_badge)
        left.addLayout(heading)
        explorer_hint = QLabel("Encuentra un proceso finalizado usando el buscador o los filtros.", objectName="wordPanelHint")
        explorer_hint.setWordWrap(True)
        left.addWidget(explorer_hint)
        self.search = QLineEdit(objectName="reportSearch")
        self.search.setPlaceholderText("Buscar por nombre, programa o periodo...")
        self.search.setClearButtonEnabled(True)
        self.search.textChanged.connect(self._apply_filters)
        left.addWidget(self.search)

        filters = QHBoxLayout()
        self.period_filter = QComboBox(objectName="reportFilter")
        self.program_filter = QComboBox(objectName="reportFilter")
        for combo in (self.period_filter, self.program_filter):
            combo.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
            combo.setMinimumContentsLength(10)
            combo.currentTextChanged.connect(self._apply_filters)
            filters.addWidget(combo, 1)
        left.addLayout(filters)
        self.report_list = QListWidget(objectName="reportList")
        self.report_list.setMinimumWidth(360)
        self.report_list.currentItemChanged.connect(self._report_selected)
        self.report_list.itemDoubleClicked.connect(lambda _: self._generate())
        left.addWidget(self.report_list, 1)
        list_actions = QHBoxLayout()
        refresh = QPushButton("Actualizar", objectName="wordAction")
        refresh.clicked.connect(self._load_completed_reports)
        browse = QPushButton("+  Seleccionar otro Excel", objectName="wordSelectAction")
        browse.clicked.connect(self._select_excel)
        list_actions.addWidget(refresh)
        list_actions.addWidget(browse)
        left.addLayout(list_actions)
        content.addWidget(explorer, 3)

        details = QFrame(objectName="wordPanel")
        right = QVBoxLayout(details)
        right.setContentsMargins(20, 16, 20, 16)
        detail_heading = QHBoxLayout()
        detail_heading.addWidget(QLabel("Resumen del documento", objectName="wordSection"))
        detail_heading.addStretch()
        self.selection_badge = QLabel("SIN SELECCIÓN", objectName="selectionBadge")
        detail_heading.addWidget(self.selection_badge)
        right.addLayout(detail_heading)
        preview_hint = QLabel("Confirma la información que se insertará en la plantilla institucional.", objectName="wordPanelHint")
        preview_hint.setWordWrap(True)
        right.addWidget(preview_hint)
        self.selection_title = QLabel("Selecciona un Excel", objectName="selectionTitle")
        self.selection_title.setWordWrap(True)
        right.addWidget(self.selection_title)
        self.detail_values = {}
        detail_grid = QGridLayout()
        fields = (("program", "PROGRAMA"), ("period", "PERIODO"), ("events", "EVENTOS REGISTRADOS"), ("file", "ARCHIVO DE ORIGEN"), ("output", "DOCUMENTO DE SALIDA"))
        for row, (key, label) in enumerate(fields):
            detail_grid.addWidget(QLabel(label, objectName="detailLabel"), row * 2, 0)
            value = QLabel("-", objectName="detailValue")
            value.setWordWrap(True)
            self.detail_values[key] = value
            detail_grid.addWidget(value, row * 2 + 1, 0)
        right.addLayout(detail_grid)
        right.addStretch()
        action_area = QFrame(objectName="wordActionArea")
        action_box = QVBoxLayout(action_area)
        action_box.setContentsMargins(12, 11, 12, 12)
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.hide()
        action_box.addWidget(self.progress)
        self.feedback = QLabel("Selecciona un Excel terminado para continuar.", objectName="excelFeedback")
        self.feedback.setWordWrap(True)
        action_box.addWidget(self.feedback)
        self.generate_button = QPushButton("Crear documento Word  →", objectName="wordPrimary")
        self.generate_button.setEnabled(False)
        self.generate_button.clicked.connect(self._generate)
        action_box.addWidget(self.generate_button)
        generated_actions = QHBoxLayout()
        self.open_button = QPushButton("Abrir Word", objectName="wordAction")
        self.pdf_button = QPushButton("Generar PDF", objectName="wordAction")
        self.open_button.clicked.connect(self._open_generated)
        self.pdf_button.clicked.connect(self._generate_pdf)
        self.open_button.hide()
        self.pdf_button.hide()
        generated_actions.addWidget(self.open_button)
        generated_actions.addWidget(self.pdf_button)
        action_box.addLayout(generated_actions)
        right.addWidget(action_area)
        content.addWidget(details, 2)
        self._load_completed_reports()

    def reset(self):
        self.search.clear()
        self.generated_path = None
        self.generated_pdf_path = None
        self.open_button.hide()
        self.pdf_button.hide()
        self._load_completed_reports()

    def _load_completed_reports(self, *_):
        self.reports = [r for r in list_completed_processes(self.user) if Path(r["workbook_path"]).is_file()]
        self._populate_filter(self.period_filter, "Todos los periodos", "period")
        self._populate_filter(self.program_filter, "Todos los programas", "program")
        self._apply_filters()

    def _populate_filter(self, combo, all_text, key):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.addItem(all_text, None)
        for value in sorted({str(r.get(key) or "Usuario") for r in self.reports}):
            combo.addItem(value, value)
        index = combo.findText(current)
        combo.setCurrentIndex(index if index >= 0 else 0)
        combo.blockSignals(False)

    def _excel_icon(self):
        """Crea un icono verde tipo Excel sin depender de un SVG externo."""
        pixmap = QPixmap(32, 32)
        pixmap.fill(Qt.transparent)

        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        # Hoja/documento
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#EAF6EF"))
        painter.drawRoundedRect(4, 2, 24, 28, 5, 5)

        # Franja Excel
        painter.setBrush(QColor("#107C41"))
        painter.drawRoundedRect(3, 7, 20, 19, 4, 4)

        # Letra X
        painter.setPen(QColor("#FFFFFF"))
        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        painter.setFont(font)
        painter.drawText(3, 7, 20, 19, Qt.AlignCenter, "X")

        painter.end()
        return QIcon(pixmap)

    def _apply_filters(self, *_):
        query = self.search.text().strip().casefold()
        period, program = self.period_filter.currentData(), self.program_filter.currentData()
        filtered = []
        for report in self.reports:
            haystack = " ".join(str(report.get(k) or "") for k in ("workbook_path", "period", "program", "level", "modality")).casefold()
            if query and query not in haystack:
                continue
            if period and report.get("period") != period:
                continue
            if program and report.get("program") != program:
                continue
            filtered.append(report)
        self.report_list.clear()
        for report in filtered:
            path = Path(report["workbook_path"])
            context = "  |  ".join(
                value for value in (
                    str(report.get("period") or ""),
                    str(report.get("level") or ""),
                    str(report.get("modality") or ""),
                ) if value
            )
            text = f'{report.get("program", "Sin programa")}\n{context}\n{path.name}'
            item = QListWidgetItem(text)
            item.setIcon(self._excel_icon())
            item.setData(Qt.UserRole, report)
            item.setSizeHint(QSize(0, 78))
            self.report_list.addItem(item)
        self.count_badge.setText(f"{len(filtered)} disponibles")
        if filtered:
            self.report_list.setCurrentRow(0)
        else:
            self._clear_selection("No hay resultados con esos filtros.")

    def _report_selected(self, current, _previous=None):
        if current:
            report = current.data(Qt.UserRole)
            self._set_selection(Path(report["workbook_path"]), report)

    def _set_selection(self, path, report):
        self.excel_path = path
        self.generated_path = None
        self.generated_pdf_path = None
        self.current_program = str(report.get("program") or "")
        self.current_period = str(report.get("period") or "")
        self.selection_badge.setText("LISTO PARA GENERAR")
        self.selection_title.setText(path.stem)
        self.detail_values["program"].setText(self.current_program or "Sin identificar")
        self.detail_values["period"].setText(self.current_period or "Sin identificar")
        self.detail_values["file"].setText(path.name)
        self.detail_values["output"].setText(self._destination().name)
        self.detail_values["events"].setText(self._read_event_total(path))
        self.generate_button.setEnabled(bool(self.current_program and self.current_period))
        self.feedback.setText("Excel listo. Revisa los datos y crea el documento Word.")
        self.open_button.hide()
        self.pdf_button.hide()

    def _read_event_total(self, path):
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook["Resumen Informe"]
                for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
                    if str(row[0] or "").strip().casefold() == "eventos totales":
                        return f"{int(float(row[1] or 0)):,}".replace(",", ".")
            finally:
                workbook.close()
        except Exception:
            return "No disponible"
        return "No disponible"

    def _clear_selection(self, message):
        self.excel_path = None
        self.selection_badge.setText("SIN SELECCIÓN")
        self.selection_title.setText("Selecciona un Excel")
        for value in self.detail_values.values():
            value.setText("-")
        self.generate_button.setEnabled(False)
        self.feedback.setText(message)

    def _select_excel(self):
        selected, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel terminado", "", "Archivos Excel (*.xlsx)")
        if not selected:
            return
        path = Path(selected)
        period = next((p for p in list_report_options("period") if p in path.stem), "")
        program = next((p for p in list_report_options("program") if p.casefold() in path.stem.replace("_", " ").casefold()), "")
        if not period or not program:
            show_error(self, "No se pudo identificar el informe", "Selecciona un Excel generado y terminado desde esta aplicacion.")
            return
        self.report_list.clearSelection()
        self._set_selection(path, {"program": program, "period": period, "owner_name": "Archivo externo"})

    def _destination(self):
        return build_word_path(self.excel_path.parent, self.current_period, self.current_program)

    def _pdf_destination(self):
        return build_pdf_path(self.excel_path.parent, self.current_period, self.current_program)

    def _generate(self):
        if not self.excel_path:
            return
        destination = self._destination()
        if destination.exists() and not ask_confirmation(self, "El informe ya existe", f"Ya existe {destination.name}.\n\nDeseas reemplazarlo?"):
            return
        self.generate_button.setEnabled(False)
        self.progress.show()
        self.feedback.setText("Generando graficos, tablas y documento institucional...")
        self._thread = QThread(self)
        self._worker = WordGenerationTask(lambda: generate_word_report(self.excel_path, destination, self.current_program, self.current_period))
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished.connect(self._generated)
        self._worker.failed.connect(self._failed)
        self._worker.finished.connect(self._thread.quit)
        self._worker.failed.connect(self._thread.quit)
        self._worker.finished.connect(self._worker.deleteLater)
        self._worker.failed.connect(self._worker.deleteLater)
        self._thread.finished.connect(self._thread_finished)
        self._thread.start()

    def _generated(self, path):
        self.generated_path = Path(path)
        self.feedback.setText(f"Informe creado correctamente: {self.generated_path.name}")
        self.open_button.show()
        self.pdf_button.show()
        show_info(self, "Informe generado", f"El documento se creo correctamente.\n\n{self.generated_path.name}", success=True)

    def _failed(self, message):
        self.feedback.setText("No fue posible crear el documento.")
        show_error(self, "Error al generar el Word", message)

    def _open_generated(self):
        if self.generated_path:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.generated_path.resolve())))

    def _generate_pdf(self):
        if not self.generated_path or not self.generated_path.is_file():
            show_error(self, "Genera primero el Word", "Primero crea el documento Word para poder generar el PDF.")
            return
        destination = self._pdf_destination()
        if destination.exists() and not ask_confirmation(self, "El PDF ya existe", f"Ya existe {destination.name}.\n\nDeseas reemplazarlo?"):
            return
        self.pdf_button.setEnabled(False)
        self.progress.show()
        self.feedback.setText("Generando PDF a partir del documento Word...")
        self._thread = QThread(self)
        self._worker = WordGenerationTask(lambda: convert_word_to_pdf(self.generated_path, destination))
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished.connect(self._pdf_generated)
        self._worker.failed.connect(self._pdf_failed)
        self._worker.finished.connect(self._thread.quit)
        self._worker.failed.connect(self._thread.quit)
        self._worker.finished.connect(self._worker.deleteLater)
        self._worker.failed.connect(self._worker.deleteLater)
        self._thread.finished.connect(self._thread_finished)
        self._thread.start()

    def _pdf_generated(self, path):
        self.generated_pdf_path = Path(path)
        self.feedback.setText(f"PDF creado correctamente: {self.generated_pdf_path.name}")
        show_info(self, "PDF generado", f"El PDF se guardo correctamente en la carpeta.\n\n{self.generated_pdf_path.name}", success=True)

    def _pdf_failed(self, message):
        self.feedback.setText("No fue posible crear el PDF.")
        show_error(self, "Error al generar el PDF", message)

    @Slot()
    def _thread_finished(self):
        thread = self._thread
        self._worker = None
        self._thread = None
        if thread:
            thread.deleteLater()
        self.progress.hide()
        self.generate_button.setEnabled(bool(self.excel_path))
        self.pdf_button.setEnabled(bool(self.generated_path))