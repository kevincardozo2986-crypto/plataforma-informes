import pandas as pd
import pytest
from openpyxl import load_workbook

from app.services.csv_service import iter_csv_chunks, prepare_original_data, read_csv_file
from app.services.excel_service import ExcelProcess, suggested_filename


def test_read_and_prepare_original_csv(tmp_path):
    csv_path = tmp_path / "moodle.csv"
    csv_path.write_text("Curso;FechaUnix;Usuario\nÁlgebra;1704067200;ana\n", encoding="utf-8")

    frame = read_csv_file(csv_path)
    prepared = prepare_original_data(frame)

    assert list(frame.columns) == ["Curso", "FechaUnix", "Usuario"]
    assert list(prepared.columns) == ["Curso", "FechaUnix", "Fecha", "Mes", "Dia", "Usuario"]
    assert str(prepared.loc[0, "Fecha"]) == "2023-12-31"
    assert prepared.loc[0, "Mes"] == 12
    assert prepared.loc[0, "Dia"] == 31


def test_excel_process_uses_one_workbook_and_limits_preview(tmp_path):
    frame = pd.DataFrame({"Curso": [f"Curso {index}" for index in range(250)]})
    process = ExcelProcess()
    process.create_original(frame)

    headers, rows, total = process.preview_sheet("Original", limit=200)
    destination = tmp_path / "resultado.xlsx"
    process.save_as(destination)

    assert process.sheet_names() == ["Original"]
    assert headers == ["Curso"]
    assert len(rows) == 200
    assert total == 250
    assert destination.is_file()


def test_finalize_excel_moves_workbook_and_removes_temporary_file(tmp_path):
    process = ExcelProcess(tmp_path / "Informe_EN_PROCESO.xlsx")
    process.create_original(pd.DataFrame({"Curso": ["Álgebra"]}))
    temporal = process.path
    destination = tmp_path / "Informe.xlsx"

    process.finalize_as(destination)

    assert not temporal.exists()
    assert destination.is_file()
    assert process.path == destination.resolve()
    assert process.sheet_names() == ["Original"]


def test_large_csv_is_written_in_chunks(tmp_path):
    csv_path = tmp_path / "grande.csv"
    csv_path.write_text("Curso;FechaUnix\n" + "\n".join(f"Curso {i};1704067200" for i in range(1200)), encoding="utf-8")
    process = ExcelProcess()

    rows, columns = process.create_original_from_chunks(
        iter_csv_chunks(csv_path, chunksize=100, prepare=True)
    )

    assert (rows, columns) == (1200, 5)
    assert process.preview_sheet("Original", limit=25)[2] == 1200


def test_suggested_filename_normalizes_program():
    assert suggested_filename("2026-1", "Ingeniería de Sistemas") == "Informe_2026-1_Ingeniería_de_Sistemas.xlsx"


def test_fechaunix_in_milliseconds_is_detected(tmp_path):
    csv_path = tmp_path / "milisegundos.csv"
    csv_path.write_text("FechaUnix\n1704067200000\n", encoding="utf-8")

    prepared = prepare_original_data(read_csv_file(csv_path))

    assert str(prepared.loc[0, "Fecha"]) == "2023-12-31"
    assert prepared.loc[0, "Mes"] == 12
    assert prepared.loc[0, "Dia"] == 31


def test_crea_tabla_docentes_con_dias_distintos_y_meses_dinamicos():
    original = pd.DataFrame(
        {
            "curso": [
                "Curso A", "Curso A", "Curso A", "Curso A", "Curso A",
                "Curso A", "Curso A", "Curso B",
            ],
            "usuario": [
                "Ana", "Ana", "Ana", "Ana", "Ana", "Ana", "Estudiante", "Luis",
            ],
            "rol": [
                "editingteacher", "editingteacher", "editingteacher",
                "editingteacher", "editingteacher", " EditingTeacher ",
                "student", "editingteacher",
            ],
            "Mes": [2, 2, 2, 3, 3, 3, 2, 3],
            "Dia": [1, 1, 2, 1, 3, 3, 9, 8],
            "Fecha": pd.to_datetime(
                [
                    "2026-02-01", "2026-02-01", "2026-02-02",
                    "2026-03-01", "2026-03-03", "2026-03-03",
                    "2026-02-09", "2026-03-08",
                ]
            ),
        }
    )
    proceso = ExcelProcess()
    proceso.create_original_from_chunks([original])

    cantidad_filas, meses = proceso.crear_tabla_docentes()

    assert cantidad_filas == 2
    assert meses == ["FEB", "MAR"]
    assert proceso.sheet_names() == ["Original", "Tabla Dinamica Docentes"]
    encabezados, filas, total = proceso.preview_sheet("Tabla Dinamica Docentes")
    assert encabezados == ["CURSO", "DOCENTE", "FEB", "MAR", "TOTAL"]
    assert total == 3
    assert filas[0] == ("Curso A", "Ana", 2, 2, 4)
    assert filas[1] == ("Curso B", "Luis", 0, 1, 1)
    assert filas[2] == ("TOTAL GENERAL", None, 1, 1.5, None)

    libro = load_workbook(proceso.path)
    hoja = libro["Tabla Dinamica Docentes"]
    assert hoja.freeze_panes == "C2"
    assert hoja["A1"].font.bold is True
    assert hoja["C2"].alignment.horizontal == "center"
    assert hoja["E2"].font.bold is True
    assert hoja["A4"].value == "TOTAL GENERAL"
    assert hoja["A4"].font.bold is True
    libro.close()


def test_tabla_docentes_reemplaza_la_misma_hoja_al_repetirse():
    original = pd.DataFrame(
        {
            "curso": ["Curso A"],
            "usuario": ["Ana"],
            "rol": ["editingteacher"],
            "Mes": [4],
            "Dia": [7],
        }
    )
    proceso = ExcelProcess()
    proceso.create_original(original)

    proceso.crear_tabla_docentes()
    proceso.crear_tabla_docentes()

    assert proceso.sheet_names() == ["Original", "Tabla Dinamica Docentes"]


def test_tabla_docentes_exige_registros_de_docentes():
    proceso = ExcelProcess()
    proceso.create_original(
        pd.DataFrame(
            {
                "curso": ["Curso A"],
                "usuario": ["Estudiante"],
                "rol": ["student"],
                "Mes": [2],
                "Dia": [1],
            }
        )
    )

    with pytest.raises(ValueError, match="editingteacher"):
        proceso.crear_tabla_docentes()


def test_crea_docentes_dg_con_promedios_y_grafica_nativa():
    original = pd.DataFrame(
        {
            "curso": ["Curso A", "Curso A", "Curso A", "Curso B"],
            "usuario": ["Ana", "Ana", "Ana", "Luis"],
            "rol": ["editingteacher"] * 4,
            "Mes": [2, 2, 3, 3],
            "Dia": [1, 2, 3, 4],
        }
    )
    proceso = ExcelProcess()
    proceso.create_original_from_chunks([original])
    proceso.crear_tabla_docentes()

    promedios = proceso.crear_grafica_docentes("Ingeniería de Sistemas", "2026-1")

    assert promedios == [("FEB", 1.0), ("MAR", 1.0)]
    assert proceso.sheet_names() == [
        "Original", "Tabla Dinamica Docentes", "Docentes DG"
    ]
    encabezados, filas, total = proceso.preview_sheet("Docentes DG")
    assert encabezados == ["MES", "PROMEDIO"]
    assert filas == [("FEB", 1), ("MAR", 1)]
    assert total == 2
    libro = load_workbook(proceso.path)
    hoja = libro["Docentes DG"]
    assert len(hoja._charts) == 1
    assert "Ingeniería de Sistemas 2026-1" in str(hoja._charts[0].title)
    libro.close()


def test_docentes_dg_se_regenera_sin_duplicar_la_hoja():
    proceso = ExcelProcess()
    proceso.create_original(
        pd.DataFrame(
            {
                "curso": ["Curso A"], "usuario": ["Ana"],
                "rol": ["editingteacher"], "Mes": [4], "Dia": [7],
            }
        )
    )
    proceso.crear_tabla_docentes()
    proceso.crear_grafica_docentes("Derecho", "2026-2")
    proceso.crear_grafica_docentes("Derecho", "2026-2")

    assert proceso.sheet_names().count("Docentes DG") == 1


def test_crea_tabla_dinamica_estudiantes_con_dias_y_usuarios_unicos():
    original = pd.DataFrame(
        {
            "curso": ["Curso A", "Curso A", "Curso A", "Curso A", "Curso B", "Curso A"],
            "idusuario": [10, 10, 11, 11, 20, 99],
            "usuario": ["Ana", "Ana", "Beto", "Beto", "Cata", "Docente"],
            "rol": ["student", "student", "student", "student", "student", "editingteacher"],
            "Mes": [2, 2, 2, 3, 3, 2],
            "Dia": [1, 1, 2, 4, 5, 9],
            "accion": ["viewed", "viewed", "created", "updated", "deleted", "viewed"],
        }
    )
    proceso = ExcelProcess()
    proceso.create_original_from_chunks([original])
    proceso.crear_tabla_docentes()
    proceso.crear_grafica_docentes("Ingeniería", "2026-1")

    cantidad, meses = proceso.crear_tabla_estudiantes("Ingeniería", "2026-1")

    assert cantidad == 2
    assert meses == ["FEB", "MAR"]
    assert proceso.sheet_names() == [
        "Original", "Tabla Dinamica Docentes", "Docentes DG",
        "Tabla Dinamica Estudiantes",
    ]
    libro = load_workbook(proceso.path)
    hoja = libro["Tabla Dinamica Estudiantes"]
    assert [hoja.cell(1, columna).value for columna in range(1, 3)] == [
        "rol", "student",
    ]
    assert [hoja.cell(2, columna).value for columna in range(1, 8)] == [
        "CURSO", "DÍAS", None, None, "ESTUDIANTES", None, None,
    ]
    assert [hoja.cell(3, columna).value for columna in range(1, 8)] == [
        None, "FEB", "MAR", "TOTAL", "FEB", "MAR", "TOTAL",
    ]
    assert hoja["D4"].value == "=SUM(B4:C4)"
    assert hoja["G4"].value == "=SUM(E4:F4)"
    assert hoja["A6"].value == "TOTAL GENERAL"
    assert hoja["I4"].value == "=MAX(D4:D5)"
    assert len(libro["Docentes DG"]._charts) == 1
    libro.close()

    libro_valores = load_workbook(proceso.path, data_only=True)
    hoja_valores = libro_valores["Tabla Dinamica Estudiantes"]
    assert [hoja_valores.cell(4, columna).value for columna in range(1, 8)] == [
        "Curso A", 2, 1, 3, 2, 1, 3,
    ]
    assert [hoja_valores.cell(5, columna).value for columna in range(1, 8)] == [
        "Curso B", 0, 1, 1, 0, 1, 1,
    ]
    assert hoja_valores["I4"].value == 3
    assert hoja_valores["B9"].value == 1
    assert hoja_valores["E9"].value == 1
    libro_valores.close()

    meses_graficas = proceso.crear_grafica_estudiantes("Ingeniería", "2026-1")
    assert meses_graficas == ["FEB", "MAR"]
    assert proceso.sheet_names() == [
        "Original", "Tabla Dinamica Docentes", "Docentes DG",
        "Tabla Dinamica Estudiantes", "Estudiantes DG", "Estudiantes DG2",
    ]
    libro = load_workbook(proceso.path)
    assert len(libro["Estudiantes DG"]._charts) == 1
    assert "estudiantes" in str(libro["Estudiantes DG"]._charts[0].title).lower()
    assert len(libro["Estudiantes DG2"]._charts) == 1
    assert "promedio de estudiantes" in str(libro["Estudiantes DG2"]._charts[0].title).lower()
    libro.close()

    cantidad_cursos, acciones = proceso.crear_tabla_actividades(
        "Ingeniería", "2026-1"
    )
    assert cantidad_cursos == 2
    assert acciones == ["created", "deleted", "updated", "viewed"]
    libro = load_workbook(proceso.path, data_only=True)
    hoja_actividades = libro["Tabla Dinamica Actividades"]
    assert [hoja_actividades.cell(4, columna).value for columna in range(1, 7)] == [
        "Etiquetas de fila", "created", "deleted", "updated", "viewed", "Total general",
    ]
    assert [hoja_actividades.cell(5, columna).value for columna in range(1, 7)] == [
        "Curso A", 1, None, 1, 3, 5,
    ]
    assert [hoja_actividades.cell(7, columna).value for columna in range(1, 7)] == [
        "Total general", 1, 1, 1, 3, 6,
    ]
    assert len(libro["Estudiantes DG"]._charts) == 1
    assert len(libro["Estudiantes DG2"]._charts) == 1
    libro.close()

    indicadores = proceso.crear_diseno_cursos("Ingeniería", "2026-1")
    assert indicadores == (0, 0, 2)
    libro = load_workbook(proceso.path, data_only=True)
    diseno = libro["Diseño de Cursos"]
    assert [diseno.cell(5, columna).value for columna in range(3, 7)] == [
        "Unificados", "Sin contenido", "Con contenido", "Total",
    ]
    assert [diseno.cell(6, columna).value for columna in range(3, 7)] == [0, 0, 2, 2]
    assert len(diseno._charts) == 1
    assert len(libro["Estudiantes DG"]._charts) == 1
    assert len(libro["Estudiantes DG2"]._charts) == 1
    libro.close()

def test_tabla_estudiantes_exige_idusuario():
    proceso = ExcelProcess()
    proceso.create_original(
        pd.DataFrame(
            {"curso": ["A"], "rol": ["student"], "Mes": [2], "Dia": [1]}
        )
    )

    with pytest.raises(ValueError, match="idusuario"):
        proceso.crear_tabla_estudiantes("Ingeniería", "2026-1")
